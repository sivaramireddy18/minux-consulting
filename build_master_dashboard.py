import os
import sys

def main():
    # Ensure openpyxl is available
    try:
        import openpyxl
    except ImportError:
        print("Installing openpyxl library inside user context...")
        import subprocess
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--user", "--break-system-packages"])
        except Exception:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--user"])
        import site
        sys.path.append(site.getusersitepackages())
        import openpyxl

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule

    # Create Workbook
    wb = openpyxl.Workbook()
    
    # Setup Sheets
    ws_dash = wb.active
    ws_dash.title = "Master Coaching Dashboard"
    ws_embedded = wb.create_sheet(title="Embedded & Linux")
    ws_c = wb.create_sheet(title="C Systems Programming")
    ws_python = wb.create_sheet(title="Python Automation")
    ws_guide = wb.create_sheet(title="Coaching Guide")

    # Ensure grid lines are visible on all sheets
    for ws in [ws_dash, ws_embedded, ws_c, ws_python, ws_guide]:
        ws.views.sheetView[0].showGridLines = True

    # -------------------------------------------------------------
    # Styling definitions
    # -------------------------------------------------------------
    FONT_FAMILY = "Segoe UI"
    
    # Fonts
    font_title = Font(name=FONT_FAMILY, size=16, bold=True, color="0F172A")
    font_section = Font(name=FONT_FAMILY, size=11, bold=True, color="1E293B")
    font_header = Font(name=FONT_FAMILY, size=9.5, bold=True, color="FFFFFF")
    font_body = Font(name=FONT_FAMILY, size=9.5, color="334155")
    font_body_bold = Font(name=FONT_FAMILY, size=9.5, bold=True, color="1E293B")
    
    font_kpi_num = Font(name=FONT_FAMILY, size=18, bold=True, color="1E3A8A")
    font_kpi_lbl = Font(name=FONT_FAMILY, size=8, bold=True, color="64748B")
    
    font_card_lbl = Font(name=FONT_FAMILY, size=9, bold=True, color="475569")
    font_card_val = Font(name=FONT_FAMILY, size=9.5, bold=True, color="0F172A")

    # Header Fills
    fill_slate = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    fill_teal = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    fill_rust = PatternFill(start_color="C2410C", end_color="C2410C", fill_type="solid")
    fill_blue = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_kpi = PatternFill(start_color="F0F9FF", end_color="F0F9FF", fill_type="solid")
    fill_card = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    fill_warn_light = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")

    # Alignments
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")
    
    # Borders
    thin_side = Side(border_style="thin", color="CBD5E1")
    border_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_double_bottom = Border(bottom=Side(style="double", color="0F172A"), top=Side(style="thin", color="CBD5E1"))
    border_kpi = Border(left=Side(style="medium", color="1E3A8A"), right=thin_side, top=thin_side, bottom=thin_side)

    # -------------------------------------------------------------
    # Shared Helper Functions
    # -------------------------------------------------------------
    def style_card_cell(ws, row, col, label, value):
        lbl_c = ws.cell(row=row, column=col, value=label)
        lbl_c.font = font_card_lbl
        lbl_c.fill = fill_card
        lbl_c.alignment = align_left
        lbl_c.border = border_thin
        
        val_c = ws.cell(row=row, column=col+1, value=value)
        val_c.font = font_card_val
        val_c.fill = fill_card
        val_c.alignment = align_left
        val_c.border = border_thin

    def setup_kpi_card_master(ws, label, formula, format_str, col_start, col_end, row_num):
        ws.merge_cells(f"{col_start}{row_num}:{col_end}{row_num}")
        lbl_cell = ws[f"{col_start}{row_num}"]
        lbl_cell.value = label
        lbl_cell.font = font_kpi_lbl
        lbl_cell.alignment = align_center
        lbl_cell.fill = fill_kpi
        lbl_cell.border = border_kpi
        
        ws.merge_cells(f"{col_start}{row_num+1}:{col_end}{row_num+1}")
        val_cell = ws[f"{col_start}{row_num+1}"]
        val_cell.value = formula
        val_cell.font = font_kpi_num
        val_cell.alignment = align_center
        val_cell.fill = fill_kpi
        val_cell.border = border_kpi
        if format_str:
            val_cell.number_format = format_str

    # -------------------------------------------------------------
    # Populate Sheet 2: Embedded & Linux (36 Weeks)
    # -------------------------------------------------------------
    headers_emb = [
        "Week", "Phase / Module", "Core Focus Area", "Status", 
        "Technical Score (1-10)", "C Code Quality", "Git Hygiene (1-5)", 
        "Hardware Safety (1-5)", "Coach's Action Plan & Remediation Notes"
    ]
    for col_idx, h in enumerate(headers_emb, 1):
        cell = ws_embedded.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_teal
        cell.alignment = align_center
        cell.border = border_thin
    ws_embedded.row_dimensions[4].height = 28

    curriculum_emb = [
        ("Week 01", "Phase 1: Foundations", "Git Workflows, Repo Discipline, and Dev Setup", "Completed", 9, "Pass", 5, 5, "Excellent repo setup and hook configurations."),
        ("Week 02", "Phase 1: Foundations", "Digital Logic, Radix Formats, Boolean Gating", "Completed", 8, "Pass", 4, 4, "Half-adder functions clean. Logic delays mapped."),
        ("Week 03", "Phase 1: Foundations", "Embedded C Basics, Variable Lifetimes, Volatile", "In Progress", 8, "Pending", 4, 5, "Working on volatile polling comparisons. Makefile clean."),
    ] + [
        (f"Week {w:02d}", "Phase 1: Foundations" if w <= 6 else "Phase 2: Bare-Metal" if w <= 12 else "Phase 3: Systems Prog" if w <= 18 else "Phase 4: RTOS Core" if w <= 24 else "Phase 5: Kernel Space" if w <= 30 else "Phase 6: Integration", "Core concepts coverage", "Not Started", "", "Pending", "", "", "") for w in range(4, 37)
    ]

    for idx, row_data in enumerate(curriculum_emb, 5):
        ws_embedded.row_dimensions[idx].height = 20
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_embedded.cell(row=idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = border_thin
            if col_idx in [2, 3, 9]:
                cell.alignment = align_left
            elif col_idx in [1, 4, 6]:
                cell.alignment = align_center
            elif col_idx in [5, 7, 8]:
                cell.alignment = align_right
            if col_idx == 1:
                cell.font = font_body_bold
        if idx % 2 == 0:
            for c in range(1, 10):
                ws_embedded.cell(row=idx, column=c).fill = fill_zebra

    # Header title
    ws_embedded.merge_cells("A1:I2")
    ws_embedded.cell(row=1, column=1, value="EMBEDDED SYSTEMS & LINUX PROGRESS TRACKER").font = font_title
    ws_embedded.cell(row=1, column=1).alignment = align_center

    # -------------------------------------------------------------
    # Populate Sheet 3: C Systems Programming (12 Weeks)
    # -------------------------------------------------------------
    headers_c = [
        "Week", "Phase / Module", "Core Focus Area", "Status", 
        "Technical Score (1-10)", "MISRA C Compliance", "Static Analysis Errors", 
        "GDB Debugging (1-5)", "Coach's Action Plan & Remediation Notes"
    ]
    for col_idx, h in enumerate(headers_c, 1):
        cell = ws_c.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_rust
        cell.alignment = align_center
        cell.border = border_thin
    ws_c.row_dimensions[4].height = 28

    curriculum_c = [
        ("Week 01", "Milestone 1: The Machine", "Compiler Mechanics & Build Systems", "In Progress", 8, "Pending", 0, 4, "Configuring auto-dependency Makefile rules."),
    ] + [
        (f"Week {w:02d}", "Milestone 1: The Machine" if w <= 3 else "Milestone 2: Memory Context" if w <= 6 else "Milestone 3: Data Boundaries" if w <= 9 else "Milestone 4: Concurrency", f"Milestone C concepts mapping W{w}", "Not Started", "", "Pending", "", "", "") for w in range(2, 13)
    ]

    for idx, row_data in enumerate(curriculum_c, 5):
        ws_c.row_dimensions[idx].height = 20
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_c.cell(row=idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = border_thin
            if col_idx in [2, 3, 9]:
                cell.alignment = align_left
            elif col_idx in [1, 4, 6]:
                cell.alignment = align_center
            elif col_idx in [5, 7, 8]:
                cell.alignment = align_right
            if col_idx == 1:
                cell.font = font_body_bold
        if idx % 2 == 0:
            for c in range(1, 10):
                ws_c.cell(row=idx, column=c).fill = fill_zebra

    ws_c.merge_cells("A1:I2")
    ws_c.cell(row=1, column=1, value="C SYSTEMS PROGRAMMING PROGRESS TRACKER").font = font_title
    ws_c.cell(row=1, column=1).alignment = align_center

    # -------------------------------------------------------------
    # Populate Sheet 4: Python Automation (12 Weeks)
    # -------------------------------------------------------------
    headers_py = [
        "Week", "Phase / Module", "Core Focus Area", "Status", 
        "Technical Score (1-10)", "Mypy Type Quality", "HIL Test Coverage %", 
        "Concurrency Safety", "Coach's Action Plan & Remediation Notes"
    ]
    for col_idx, h in enumerate(headers_py, 1):
        cell = ws_python.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_blue
        cell.alignment = align_center
        cell.border = border_thin
    ws_python.row_dimensions[4].height = 28

    curriculum_py = [
        ("Week 01", "Milestone 1: Embedded Py", "Host Automation & CLI Wrapping", "In Progress", 9, "Pass", 0.95, "Pass", "Clean type safety setup and robust subprocess timeouts."),
    ] + [
        (f"Week {w:02d}", "Milestone 1: Embedded Py" if w <= 3 else "Milestone 2: HIL Testing" if w <= 6 else "Milestone 3: Binary Audits" if w <= 9 else "Milestone 4: Enterprise CI/CD", f"Milestone Py concepts mapping W{w}", "Not Started", "", "Pending", "", "Pending", "") for w in range(2, 13)
    ]

    for idx, row_data in enumerate(curriculum_py, 5):
        ws_python.row_dimensions[idx].height = 20
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_python.cell(row=idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = border_thin
            if col_idx in [2, 3, 9]:
                cell.alignment = align_left
            elif col_idx in [1, 4, 6, 8]:
                cell.alignment = align_center
            elif col_idx in [5, 7]:
                cell.alignment = align_right
            if col_idx == 7 and type(val) in [float, int]:
                cell.number_format = "0%"
            if col_idx == 1:
                cell.font = font_body_bold
        if idx % 2 == 0:
            for c in range(1, 10):
                ws_python.cell(row=idx, column=c).fill = fill_zebra

    ws_python.merge_cells("A1:I2")
    ws_python.cell(row=1, column=1, value="PYTHON AUTOMATION FRAMEWORKS PROGRESS TRACKER").font = font_title
    ws_python.cell(row=1, column=1).alignment = align_center

    # -------------------------------------------------------------
    # Populate Sheet 5: Coaching Guide (Reference Manual)
    # -------------------------------------------------------------
    ws_guide.merge_cells("A1:G2")
    guide_title = ws_guide.cell(row=1, column=1, value="COACHING MANUAL & METRIC EVALUATION STANDARDS")
    guide_title.font = font_title
    guide_title.fill = fill_slate
    guide_title.alignment = align_center
    ws_guide.row_dimensions[1].height = 20
    ws_guide.row_dimensions[2].height = 20

    # Section 1: Scoring Rubric Table
    ws_guide.cell(row=4, column=1, value="1. GRANULAR SCORE CARD RUBRIC").font = font_section
    
    rubric_headers = ["Metric Category", "High Score (9-10 / 5)", "Nominal Pass (7-8 / 3-4)", "Automatic Failure (0-6 / 1-2)"]
    for col_idx, r_h in enumerate(rubric_headers, 1):
        cell = ws_guide.cell(row=5, column=col_idx, value=r_h)
        cell.font = font_header
        cell.fill = fill_slate
        cell.alignment = align_center
        cell.border = border_thin
    ws_guide.row_dimensions[5].height = 24

    rubrics_data = [
        ("Technical Score (1-10)", "Perfect algorithmic logic, optimized memory maps, and completes the Stretch Task.", "Correct dynamic execution, but minor optimizations or styling can be refined.", "Functional compile errors, logical bugs, lack of unit tests, or core requirement breach."),
        ("Code Quality (Pass/Fail)", "Pass: Zero warnings under static analysis (Cppcheck) and 100% clean Valgrind run (0 bytes leaked).", "N/A - Pass is absolute.", "Fail: Any compile warnings, static checker errors, or memory leaks/Use-After-Frees."),
        ("Git Hygiene (1-5)", "Daily atomic commits, clear imperative logs (feat:), and strict .gitignore exclusions.", "Consistent commits, but logs are generic or commits are too large.", "Pushing raw binaries (.o, .pyc, ELFs) or pushing directly to main without PR processes."),
        ("Hardware Safety (1-5)", "Neat wiring, ESD safety observed, voltage meters verified BEFORE powering up, clean emergency de-energize.", "Correct wiring but chaotic layout, minor floating input hazards.", "Wiring hardware while target MCU is powered on, triggering short circuits, raw rail overloads."),
        ("Mypy Type Safety (Pass/Fail)", "Pass: mypy --strict passes with 0 alerts, 0 type omissions, and 0 dynamic Any casts.", "N/A - Pass is absolute.", "Fail: Any type validator warnings, untyped function arguments, or parser checks failure.")
    ]

    for idx, row_data in enumerate(rubrics_data, 6):
        ws_guide.row_dimensions[idx].height = 24
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_guide.cell(row=idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = border_thin
            if col_idx == 1:
                cell.font = font_body_bold
                cell.alignment = align_left
            else:
                cell.alignment = align_left
        if idx % 2 == 1:
            for c in range(1, 5):
                ws_guide.cell(row=idx, column=c).fill = fill_zebra

    # Section 2: Weekly Review Checklist
    ws_guide.cell(row=12, column=1, value="2. WEEKLY EVALUATION CHECKLIST FOR THE COACH").font = font_section
    
    checklist_steps = [
        ("Step 1: Code Quality Audit", "For C: Run static analysis `cppcheck --enable=all src/`. For Python: Run type auditor `mypy --strict src/`. Ensure 0 warnings are returned."),
        ("Step 2: Dynamic Leak Audits", "For C: Run compiled test suite under Valgrind Memcheck: `valgrind --leak-check=full ./test_bin`. Verify 0 leaks exist."),
        ("Step 3: Concurrency Safety", "For Python: Run HIL suites under Helgrind: `valgrind --tool=helgrind python -m pytest`. Ensure 0 race warnings exist."),
        ("Step 4: Update Excel & Alert", "Record grades in corresponding sheets. If any check fails, toggle quality to 'Fail'—this automatically flags a global red alert on Tab 1.")
    ]

    for idx, (step, desc) in enumerate(checklist_steps, 13):
        ws_guide.row_dimensions[idx].height = 20
        cell_step = ws_guide.cell(row=idx, column=1, value=step)
        cell_step.font = font_body_bold
        cell_step.border = border_thin
        cell_step.alignment = align_left
        
        cell_desc = ws_guide.cell(row=idx, column=2, value=desc)
        cell_desc.font = font_body
        cell_desc.border = border_thin
        cell_desc.alignment = align_left
        ws_guide.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=4)

    # Section 3: Ideal Data Logging Reference
    ws_guide.cell(row=18, column=1, value="3. IDEAL DATA LOGGING EXAMPLES (MEASURABLE RESULTS)").font = font_section
    
    example_headers = ["Track Module", "Status", "Tech", "Quality Grade", "Safe / Git Score", "Action Plan / Remediation Notes Example"]
    for col_idx, ex_h in enumerate(example_headers, 1):
        cell = ws_guide.cell(row=19, column=col_idx, value=ex_h)
        cell.font = font_header
        cell.fill = fill_slate
        cell.alignment = align_center
        cell.border = border_thin
    ws_guide.row_dimensions[19].height = 24

    examples_data = [
        ("C Week 01: Build Systems", "Completed", 9, "Pass (MISRA)", 5, "Matrix library compiles cleanly. Incremental Makefile dependency tracking verified (.d files). Symbol nm audits pass."),
        ("C Week 02: CPU Realities", "In Progress", 8, "Pending", 4, "Working on float coordinates parser. Reminded student to replace direct float equality checks with epsilon comparisons."),
        ("Py Week 01: Subprocesses", "Completed", 9, "Pass (Mypy)", "95% Cov", "Flashing automation tool is fully operational. Successfully captures toolchain stderr loops. TimeoutExpired limits are locked.")
    ]

    for idx, row_data in enumerate(examples_data, 20):
        ws_guide.row_dimensions[idx].height = 24
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_guide.cell(row=idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = border_thin
            if col_idx == 1:
                cell.font = font_body_bold
            if col_idx in [1, 6]:
                cell.alignment = align_left
            elif col_idx in [2, 4]:
                cell.alignment = align_center
            else:
                cell.alignment = align_right
        if idx % 2 == 1:
            for c in range(1, 7):
                ws_guide.cell(row=idx, column=c).fill = fill_zebra

    # Adjust guide columns
    ws_guide.column_dimensions["A"].width = 28
    ws_guide.column_dimensions["B"].width = 30
    ws_guide.column_dimensions["C"].width = 30
    ws_guide.column_dimensions["D"].width = 30
    ws_guide.column_dimensions["E"].width = 18
    ws_guide.column_dimensions["F"].width = 48

    # -------------------------------------------------------------
    # Populate Sheet 1: Master Coaching Dashboard
    # -------------------------------------------------------------
    # Student Profile Card
    ws_dash.cell(row=4, column=1, value="STUDENT PROFILE & GLOBAL ALERTS").font = font_section
    
    style_card_cell(ws_dash, 5, 1, "Student Name:", "Sivaramireddy")
    style_card_cell(ws_dash, 6, 1, "Assigned Mentor:", "Senior Systems Coach")
    style_card_cell(ws_dash, 5, 3, "Start Date:", "2026-06-01")
    style_card_cell(ws_dash, 6, 3, "Program Status:", "Active Learning")
    
    # Global Alert: checks for FAIL in F columns (Code Quality, MISRA, Mypy) across trackers
    global_alert_formula = (
        "=IF(OR("
        "COUNTIF('Embedded & Linux'!F5:F40, \"Fail\")>0, "
        "COUNTIF('C Systems Programming'!F5:F16, \"Fail\")>0, "
        "COUNTIF('Python Automation'!F5:F16, \"Fail\")>0"
        "), \"CRITICAL: Fail Grade Alert\", \"Nominal - On Track\")"
    )
    style_card_cell(ws_dash, 5, 5, "Global Alert Status:", global_alert_formula)
    style_card_cell(ws_dash, 6, 5, "Total Active Tracks:", "3 Active Tracks")
    
    ws_dash.row_dimensions[5].height = 20
    ws_dash.row_dimensions[6].height = 20

    # 3. Dynamic KPI Blocks (Master Layout)
    ws_dash.cell(row=8, column=1, value="TRI-DISCIPLINE PERFORMANCE TRACKING WIDGETS").font = font_section
    
    # Row 9-10: Track 1: Embedded Systems & Linux
    ws_dash.cell(row=9, column=1, value="Track 1: Embedded Systems").font = font_body_bold
    ws_dash.cell(row=9, column=1).alignment = align_left
    setup_kpi_card_master(ws_dash, "COMPLETION %", "=COUNTIF('Embedded & Linux'!D5:D40, \"Completed\")/36", "0.0%", "B", "B", 9)
    setup_kpi_card_master(ws_dash, "AVG TECH SCORE", "=AVERAGE('Embedded & Linux'!E5:E40)", "0.0", "C", "C", 9)
    setup_kpi_card_master(ws_dash, "GIT HYGIENE (1-5)", "=AVERAGE('Embedded & Linux'!G5:G40)", "0.0", "D", "D", 9)
    setup_kpi_card_master(ws_dash, "SAFETY CHECK (1-5)", "=AVERAGE('Embedded & Linux'!H5:H40)", "0.0", "E", "E", 9)
    
    ws_dash.row_dimensions[9].height = 16
    ws_dash.row_dimensions[10].height = 24

    # Row 12-13: Track 2: C Systems Programming
    ws_dash.cell(row=12, column=1, value="Track 2: C Programming").font = font_body_bold
    ws_dash.cell(row=12, column=1).alignment = align_left
    setup_kpi_card_master(ws_dash, "COMPLETION %", "=COUNTIF('C Systems Programming'!D5:D16, \"Completed\")/12", "0.0%", "B", "B", 12)
    setup_kpi_card_master(ws_dash, "AVG TECH SCORE", "=AVERAGE('C Systems Programming'!E5:E16)", "0.0", "C", "C", 12)
    setup_kpi_card_master(ws_dash, "STATIC ERRORS (AVG)", "=AVERAGE('C Systems Programming'!G5:G16)", "0", "D", "D", 12)
    setup_kpi_card_master(ws_dash, "GDB SCORE (1-5)", "=AVERAGE('C Systems Programming'!H5:H16)", "0.0", "E", "E", 12)
    
    ws_dash.row_dimensions[12].height = 16
    ws_dash.row_dimensions[13].height = 24

    # Row 15-16: Track 3: Python Automation
    ws_dash.cell(row=15, column=1, value="Track 3: Python HIL").font = font_body_bold
    ws_dash.cell(row=15, column=1).alignment = align_left
    setup_kpi_card_master(ws_dash, "COMPLETION %", "=COUNTIF('Python Automation'!D5:D16, \"Completed\")/12", "0.0%", "B", "B", 15)
    setup_kpi_card_master(ws_dash, "AVG TECH SCORE", "=AVERAGE('Python Automation'!E5:E16)", "0.0", "C", "C", 15)
    setup_kpi_card_master(ws_dash, "HIL TEST COVERAGE %", "=AVERAGE('Python Automation'!G5:G16)", "0.0%", "D", "D", 15)
    setup_kpi_card_master(ws_dash, "CONCURRENCY (PASS RATE)", "=COUNTIF('Python Automation'!H5:H16, \"Pass\")/12", "0.0%", "E", "E", 15)
    
    ws_dash.row_dimensions[15].height = 16
    ws_dash.row_dimensions[16].height = 24

    # 4. Global Program Consolidation (Row 18 to 25)
    ws_dash.cell(row=18, column=1, value="CONSOLIDATED CURRICULUM SYNOPSIS").font = font_section
    
    synopsis_headers = ["Training Discipline Track", "Syllabus Size", "Active Progress Status", "Calculated Track Score", "Current Graduation Gate Target"]
    for col_idx, sh in enumerate(synopsis_headers, 1):
        cell = ws_dash.cell(row=19, column=col_idx, value=sh)
        cell.font = font_header
        cell.fill = fill_slate
        cell.alignment = align_center
        cell.border = border_thin
    ws_dash.row_dimensions[19].height = 24

    synopsis_rows = [
        ("Track 1: Embedded Systems & Linux", "36 Weeks", "=IF(B10=1, \"Completed\", \"Active\")", "=AVERAGE('Embedded & Linux'!E5:E40)", "Capstone IIoT Gateway integration"),
        ("Track 2: Professional C Programming", "12 Weeks", "=IF(B13=1, \"Completed\", \"Active\")", "=AVERAGE('C Systems Programming'!E5:E16)", "Thread-Safe Systems Shell"),
        ("Track 3: Python Automation Frameworks", "12 Weeks", "=IF(B16=1, \"Completed\", \"Active\")", "=AVERAGE('Python Automation'!E5:E16)", "Enterprise HIL Framework Build")
    ]

    for idx, row_data in enumerate(synopsis_rows, 20):
        ws_dash.row_dimensions[idx].height = 20
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_dash.cell(row=idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = border_thin
            if col_idx in [1, 5]:
                cell.alignment = align_left
            elif col_idx in [2, 3]:
                cell.alignment = align_center
            elif col_idx == 4:
                cell.alignment = align_right
                cell.number_format = "0.0"
                cell.font = font_body_bold
        if idx % 2 == 1:
            for c in range(1, 6):
                ws_dash.cell(row=idx, column=c).fill = fill_zebra

    # 5. Coach's Intervention and Strategic Actions Area (Row 24 to 28)
    ws_dash.cell(row=24, column=1, value="COACH'S INTERVENTION PLAN & STRATEGIC ACTION LOG").font = font_section
    ws_dash.merge_cells("A25:G25")
    act_hdr = ws_dash.cell(row=25, column=1, value="Active Blockages & Mentor Directive Actions")
    act_hdr.font = font_header
    act_hdr.fill = fill_slate
    act_hdr.alignment = align_left
    act_hdr.border = border_thin

    ws_dash.merge_cells("A26:G27")
    act_body = ws_dash.cell(row=26, column=1, value="Sivaramireddy is executing exceptionally. Embedded Weeks 1-2 and C/Python Week 1 are in progress. Focus on making sure compiler flags are completely mastered this week. Ensure Mypy strict checks pass without 'Any' type bypasses inside the Python scripts.")
    act_body.font = font_body
    act_body.alignment = align_left
    act_body.border = border_thin

    ws_dash.row_dimensions[24].height = 20
    ws_dash.row_dimensions[25].height = 22
    ws_dash.row_dimensions[26].height = 20
    ws_dash.row_dimensions[27].height = 20

    # Adjust column widths dynamically for all sheets
    for ws in [ws_dash, ws_embedded, ws_c, ws_python]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row in [1, 2]: continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    # Set customized column overrides for cockpit page readability
    ws_dash.column_dimensions["A"].width = 38
    ws_dash.column_dimensions["B"].width = 18
    ws_dash.column_dimensions["C"].width = 18
    ws_dash.column_dimensions["D"].width = 18
    ws_dash.column_dimensions["E"].width = 38

    # -------------------------------------------------------------
    # Dropdowns and Data Validations
    # -------------------------------------------------------------
    from openpyxl.worksheet.datavalidation import DataValidation
    
    # Status validation list
    dv_status = DataValidation(type="list", formula1='"Not Started,In Progress,Completed"', allow_blank=True)
    # Quality validation list
    dv_quality = DataValidation(type="list", formula1='"Pass,Fail,Pending"', allow_blank=True)
    
    # Add validations to trackers sheets
    for ws in [ws_embedded, ws_c, ws_python]:
        ws.add_data_validation(dv_status)
        ws.add_data_validation(dv_quality)

    dv_status.add("D5:D40") # Embedded status
    dv_quality.add("F5:F40") # Embedded quality
    
    dv_status.add("D5:D16") # C status
    dv_quality.add("F5:F16") # C quality
    
    dv_status.add("D5:D16") # Python status
    dv_quality.add("F5:F16") # Python quality
    dv_quality.add("H5:H16") # Python concurrency safety quality

    # -------------------------------------------------------------
    # Conditional Formatting Rules
    # -------------------------------------------------------------
    fill_green = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    font_green = Font(name=FONT_FAMILY, size=9.5, bold=True, color="166534")
    
    fill_yellow = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
    font_yellow = Font(name=FONT_FAMILY, size=9.5, bold=True, color="854D0E")

    fill_gray = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    font_gray = Font(name=FONT_FAMILY, size=9.5, bold=True, color="475569")

    fill_fail = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    font_fail = Font(name=FONT_FAMILY, size=9.5, bold=True, color="991B1B")

    # Apply validations styling
    for ws in [ws_embedded, ws_c, ws_python]:
        max_row = "40" if ws == ws_embedded else "16"
        # Status Column
        ws.conditional_formatting.add(f"D5:D{max_row}", CellIsRule(operator="equal", formula=['"Completed"'], fill=fill_green, font=font_green))
        ws.conditional_formatting.add(f"D5:D{max_row}", CellIsRule(operator="equal", formula=['"In Progress"'], fill=fill_yellow, font=font_yellow))
        ws.conditional_formatting.add(f"D5:D{max_row}", CellIsRule(operator="equal", formula=['"Not Started"'], fill=fill_gray, font=font_gray))
        
        # Quality Column
        ws.conditional_formatting.add(f"F5:F{max_row}", CellIsRule(operator="equal", formula=['"Pass"'], fill=fill_green, font=font_green))
        ws.conditional_formatting.add(f"F5:F{max_row}", CellIsRule(operator="equal", formula=['"Fail"'], fill=fill_fail, font=font_fail))
        ws.conditional_formatting.add(f"F5:F{max_row}", CellIsRule(operator="equal", formula=['"Pending"'], fill=fill_gray, font=font_gray))

    # Python Concurrency Column
    ws_python.conditional_formatting.add("H5:H16", CellIsRule(operator="equal", formula=['"Pass"'], fill=fill_green, font=font_green))
    ws_python.conditional_formatting.add("H5:H16", CellIsRule(operator="equal", formula=['"Fail"'], fill=fill_fail, font=font_fail))
    ws_python.conditional_formatting.add("H5:H16", CellIsRule(operator="equal", formula=['"Pending"'], fill=fill_gray, font=font_gray))

    # Global Alert Status color in Dashboard
    ws_dash.conditional_formatting.add("F5", CellIsRule(operator="equal", formula=['"CRITICAL: Fail Grade Alert"'], fill=fill_fail, font=font_fail))
    ws_dash.conditional_formatting.add("F5", CellIsRule(operator="equal", formula=['"Nominal - On Track"'], fill=fill_green, font=font_green))

    # Save Output
    out_path = "/home/siva/sivaramireddy/Project1/curriculum_progress_tracker.xlsx"
    wb.save(out_path)
    print(f"SUCCESS: Unified Multi-Track Coaching Dashboard compiled successfully at: {out_path}")

if __name__ == "__main__":
    main()
