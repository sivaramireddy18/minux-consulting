import os
import sys

def main():
    # Ensure openpyxl is available
    try:
        import openpyxl
    except ImportError:
        print("Installing openpyxl library inside user context...")
        import subprocess
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--user", "--break-system-packages"])
        except Exception:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--user"])
        import site
        sys.path.append(site.getusersitepackages())
        import openpyxl

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule

    # Create Workbook
    wb = openpyxl.Workbook()
    
    # Setup Sheets
    ws_dash = wb.active
    ws_dash.title = "Coaching Dashboard"
    ws_tasks = wb.create_sheet(title="Weekly Progress Tracker")

    # Ensure grid lines are visible
    ws_dash.views.sheetView[0].showGridLines = True
    ws_tasks.views.sheetView[0].showGridLines = True

    # -------------------------------------------------------------
    # Premium Colors & Styling Definitions (Ocean Teal & Deep Slate)
    # -------------------------------------------------------------
    FONT_FAMILY = "Segoe UI"
    
    # Fonts
    font_title = Font(name=FONT_FAMILY, size=16, bold=True, color="0F172A")
    font_section = Font(name=FONT_FAMILY, size=12, bold=True, color="0F766E")
    font_header = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
    font_body = Font(name=FONT_FAMILY, size=10, color="334155")
    font_body_bold = Font(name=FONT_FAMILY, size=10, bold=True, color="1E293B")
    font_kpi_num = Font(name=FONT_FAMILY, size=18, bold=True, color="0F766E")
    font_kpi_lbl = Font(name=FONT_FAMILY, size=8.5, bold=True, color="64748B")
    font_card_lbl = Font(name=FONT_FAMILY, size=9, bold=True, color="475569")
    font_card_val = Font(name=FONT_FAMILY, size=10, bold=True, color="0F172A")

    # Fills
    fill_primary = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid") # Ocean Teal
    fill_accent = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Deep Slate
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Light Slate
    fill_kpi = PatternFill(start_color="F0FDFA", end_color="F0FDFA", fill_type="solid") # Very Light Teal
    fill_card = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid") # Light Gray Card
    fill_warn_light = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid") # Light Red Warning

    # Alignments
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")
    
    # Borders
    thin_border_side = Side(border_style="thin", color="CBD5E1")
    border_thin = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_kpi = Border(left=Side(style="medium", color="0F766E"), right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_double_bottom = Border(bottom=Side(style="double", color="0F172A"), top=Side(style="thin", color="CBD5E1"))

    # -------------------------------------------------------------
    # Populate Sheet 2: Weekly Progress Tracker (Coach View)
    # -------------------------------------------------------------
    headers = [
        "Week", "Phase / Module", "Core Focus Area", "Status", 
        "Technical Score (1-10)", "C Code Quality", "Git Hygiene (1-5)", 
        "Hardware Safety (1-5)", "Coach's Action Plan & Remediation Notes"
    ]
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws_tasks.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = align_center
        cell.border = border_thin
    
    ws_tasks.row_dimensions[4].height = 28

    curriculum = [
        ("Week 01", "Phase 1: Foundations", "Git Workflows, Repo Discipline, and Dev Setup", "Completed", 9, "Pass", 5, 5, "Excellent repo setup and hook configurations."),
        ("Week 02", "Phase 1: Foundations", "Digital Logic, Radix Formats, Boolean Gating", "Completed", 8, "Pass", 4, 4, "Half-adder functions clean. Logic delays mapped."),
        ("Week 03", "Phase 1: Foundations", "Embedded C Basics, Variable Lifetimes, Volatile", "In Progress", 8, "Pending", 4, 5, "Working on volatile polling comparisons. Makefile clean."),
        ("Week 04", "Phase 1: Foundations", "Pointers, App Memory Maps, Memory Allocators", "Not Started", "", "Pending", "", "", ""),
        ("Week 05", "Phase 1: Foundations", "Bitwise Math, Bitfields, Registers MMIO", "Not Started", "", "Pending", "", "", ""),
        ("Week 06", "Phase 1: Foundations", "Structures Packing, Alignment, Serialization", "Not Started", "", "Pending", "", "", ""),
        
        ("Week 07", "Phase 2: Bare-Metal", "ARM Cortex-M Core, Linker Scripts, Vectors", "Not Started", "", "Pending", "", "", ""),
        ("Week 08", "Phase 2: Bare-Metal", "Bare-Metal Clocks Gating & GPIO Registers", "Not Started", "", "Pending", "", "", ""),
        ("Week 09", "Phase 2: Bare-Metal", "Interrupts Controller (NVIC), EXTI pin routing", "Not Started", "", "Pending", "", "", ""),
        ("Week 10", "Phase 2: Bare-Metal", "SysTick Counter, General TIM interrupts, ARR/PSC", "Not Started", "", "Pending", "", "", ""),
        ("Week 11", "Phase 2: Bare-Metal", "Communication Protocols I: UART Serial Drivers", "Not Started", "", "Pending", "", "", ""),
        ("Week 12", "Phase 2: Bare-Metal", "Communication Protocols II: I2C & SPI Registers", "Not Started", "", "Pending", "", "", ""),
        
        ("Week 13", "Phase 3: Systems Prog", "Linux OS CLI Command Pipelines, FHS, Perms", "Not Started", "", "Pending", "", "", ""),
        ("Week 14", "Phase 3: Systems Prog", "GCC compiler stages, Makefile dependencies, ELF", "Not Started", "", "Pending", "", "", ""),
        ("Week 15", "Phase 3: Systems Prog", "Low-level system I/O, ioctl, memory-mapping", "Not Started", "", "Pending", "", "", ""),
        ("Week 16", "Phase 3: Systems Prog", "Process Control, fork, exec, wait, Shared Memory", "Not Started", "", "Pending", "", "", ""),
        ("Week 17", "Phase 3: Systems Prog", "POSIX threads, mutex, condition variables", "Not Started", "", "Pending", "", "", ""),
        ("Week 18", "Phase 3: Systems Prog", "Bash Automation scripts, sed stream, awk parsing", "Not Started", "", "Pending", "", "", ""),
        
        ("Week 19", "Phase 4: RTOS Core", "RTOS Paradigm, FreeRTOS ports, TCB structure", "Not Started", "", "Pending", "", "", ""),
        ("Week 20", "Phase 4: RTOS Core", "Task state machine, preemptive scheduler, PendSV", "Not Started", "", "Pending", "", "", ""),
        ("Week 21", "Phase 4: RTOS Core", "Binary/Counting Semaphores, Mutexes, Inversion", "Not Started", "", "Pending", "", "", ""),
        ("Week 22", "Phase 4: RTOS Core", "ITC, Message Queues by value, Event Groups", "Not Started", "", "Pending", "", "", ""),
        ("Week 23", "Phase 4: RTOS Core", "Memory schemes (heap_1 to 5), stack overflow", "Not Started", "", "Pending", "", "", ""),
        ("Week 24", "Phase 4: RTOS Core", "Software timers, ISR FromISR APIs, deferred", "Not Started", "", "Pending", "", "", ""),
        
        ("Week 25", "Phase 5: Kernel Space", "Embedded Linux bootloader, SPL, U-Boot CLI", "Not Started", "", "Pending", "", "", ""),
        ("Week 26", "Phase 5: Kernel Space", "Cross-compilation, Kernel source compile, boot", "Not Started", "", "Pending", "", "", ""),
        ("Week 27", "Phase 5: Kernel Space", "Device Tree DTS/DTSI node maps, compatible", "Not Started", "", "Pending", "", "", ""),
        ("Week 28", "Phase 5: Kernel Space", "Out-of-tree loadable module (LKM), parameters", "Not Started", "", "Pending", "", "", ""),
        ("Week 29", "Phase 5: Kernel Space", "Character Device registration, cdev, file_ops", "Not Started", "", "Pending", "", "", ""),
        ("Week 30", "Phase 5: Kernel Space", "remap virtual kernel map, legacy GPIO, IRQs", "Not Started", "", "Pending", "", "", ""),
        
        ("Week 31", "Phase 6: Integration", "Rootfs generation Buildroot, overlays, packages", "Not Started", "", "Pending", "", "", ""),
        ("Week 32", "Phase 6: Integration", "Yocto Poky layers, BitBake compile, recipes", "Not Started", "", "Pending", "", "", ""),
        ("Week 33", "Phase 6: Integration", "Analog-to-digital (ADC), register PWM control", "Not Started", "", "Pending", "", "", ""),
        ("Week 34", "Phase 6: Integration", "POSIX Sockets, TCP/IP concurrent multi-thread", "Not Started", "", "Pending", "", "", ""),
        ("Week 35", "Phase 6: Integration", "System profiling, gdbserver, valgrind, perf", "Not Started", "", "Pending", "", "", ""),
        ("Week 36", "Phase 6: Integration", "Capstone IoT Gateway integration, Yocto boot", "Not Started", "", "Pending", "", "", ""),
    ]

    for idx, row_data in enumerate(curriculum, 5):
        ws_tasks.row_dimensions[idx].height = 20
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_tasks.cell(row=idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = border_thin
            
            # Alignments
            if col_idx in [2, 3, 9]:
                cell.alignment = align_left
            elif col_idx in [1, 4, 6]:
                cell.alignment = align_center
            elif col_idx in [5, 7, 8]:
                cell.alignment = align_right

            # Highlight week
            if col_idx == 1:
                cell.font = font_body_bold
        
        # Zebra striping
        if idx % 2 == 0:
            for c in range(1, 10):
                ws_tasks.cell(row=idx, column=c).fill = fill_zebra

    # Dropdowns for Status and Code Quality
    from openpyxl.worksheet.datavalidation import DataValidation
    dv_status = DataValidation(type="list", formula1='"Not Started,In Progress,Completed"', allow_blank=True)
    dv_quality = DataValidation(type="list", formula1='"Pass,Fail,Pending"', allow_blank=True)
    
    ws_tasks.add_data_validation(dv_status)
    ws_tasks.add_data_validation(dv_quality)
    
    dv_status.add("D5:D40")
    dv_quality.add("F5:F40")

    # Add header in Sheet 2
    ws_tasks.merge_cells("A1:I2")
    ws_tasks.cell(row=1, column=1, value="STUDENT WEEKLY ASSESSMENT RECORD").font = font_title
    ws_tasks.cell(row=1, column=1).alignment = align_center

    # Column widths for Sheet 2
    for col in ws_tasks.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2]: continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws_tasks.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # -------------------------------------------------------------
    # Populate Sheet 1: Coaching Dashboard
    # -------------------------------------------------------------
    # 1. Title Banner
    ws_dash.merge_cells("A1:G2")
    title_cell = ws_dash.cell(row=1, column=1, value="EMBEDDED SYSTEMS & LINUX COACHING DASHBOARD")
    title_cell.font = font_title
    title_cell.alignment = align_center
    
    ws_dash.row_dimensions[1].height = 20
    ws_dash.row_dimensions[2].height = 20

    # 2. Student Profile Card (Rows 4 to 6)
    def style_card_cell(ws, row, col, label, value):
        lbl_c = ws.cell(row=row, column=col, value=label)
        lbl_c.font = font_card_lbl
        lbl_c.fill = fill_card
        lbl_c.alignment = align_left
        lbl_c.border = border_thin
        
        val_c = ws.cell(row=row, column=col+1, value=value)
        val_c.font = font_card_val
        val_c.fill = fill_card
        val_c.alignment = align_left
        val_c.border = border_thin

    ws_dash.cell(row=4, column=1, value="STUDENT PROFILE").font = font_section
    style_card_cell(ws_dash, 5, 1, "Student Name:", "Sivaramireddy")
    style_card_cell(ws_dash, 6, 1, "Assigned Mentor:", "Senior Systems Coach")
    style_card_cell(ws_dash, 5, 3, "Start Date:", "2026-06-01")
    style_card_cell(ws_dash, 6, 3, "Graduation Target:", "2027-02-05")
    style_card_cell(ws_dash, 5, 5, "Program Status:", "On Track")
    
    # Target Status Formula:
    # If any week is failed, or behind schedule -> "At Risk", else "On Track"
    style_card_cell(ws_dash, 6, 5, "Alerts/Blockers:", "=IF(COUNTIF('Weekly Progress Tracker'!F5:F40, \"Fail\")>0, \"CRITICAL: Fail Grade\", \"None\")")

    ws_dash.row_dimensions[5].height = 20
    ws_dash.row_dimensions[6].height = 20

    # 3. Coach KPI Metrics (Row 9 to 10)
    ws_dash.cell(row=8, column=1, value="PERFORMANCE METRICS OVERVIEW").font = font_section
    
    def setup_kpi_card(ws, label, formula, col_start, col_end):
        ws.merge_cells(f"{col_start}9:{col_end}9")
        lbl_cell = ws[f"{col_start}9"]
        lbl_cell.value = label
        lbl_cell.font = font_kpi_lbl
        lbl_cell.alignment = align_center
        lbl_cell.fill = fill_kpi
        lbl_cell.border = border_kpi
        
        ws.merge_cells(f"{col_start}10:{col_end}10")
        val_cell = ws[f"{col_start}10"]
        val_cell.value = formula
        val_cell.font = font_kpi_num
        val_cell.alignment = align_center
        val_cell.fill = fill_kpi
        val_cell.border = border_kpi

    setup_kpi_card(ws_dash, "CURRICULUM WEEKS", 36, "A", "A")
    setup_kpi_card(ws_dash, "WEEKS COMPLETED", "=COUNTIF('Weekly Progress Tracker'!D5:D40, \"Completed\")", "B", "B")
    setup_kpi_card(ws_dash, "IN PROGRESS", "=COUNTIF('Weekly Progress Tracker'!D5:D40, \"In Progress\")", "C", "C")
    setup_kpi_card(ws_dash, "NOT STARTED", "=COUNTIF('Weekly Progress Tracker'!D5:D40, \"Not Started\")", "D", "D")
    
    setup_kpi_card(ws_dash, "AVG TECH SCORE (1-10)", "=AVERAGE('Weekly Progress Tracker'!E5:E40)", "E", "E")
    ws_dash["E10"].number_format = "0.0"
    
    setup_kpi_card(ws_dash, "GIT DISCIPLINE (1-5)", "=AVERAGE('Weekly Progress Tracker'!G5:G40)", "F", "F")
    ws_dash["F10"].number_format = "0.0"

    setup_kpi_card(ws_dash, "PROGRAM COMPLETION %", "=B10/A10", "G", "G")
    ws_dash["G10"].number_format = "0.0%"

    ws_dash.row_dimensions[9].height = 16
    ws_dash.row_dimensions[10].height = 30

    # 4. Phase Summary Matrix Table (Row 13 to 20)
    ws_dash.cell(row=12, column=1, value="PHASE-WISE PERFORMANCE SUMMARY").font = font_section
    
    summary_headers = ["Phase / Module Name", "Weeks Included", "Status", "Completion", "Milestone Target Gate"]
    for col_idx, sh in enumerate(summary_headers, 1):
        cell = ws_dash.cell(row=13, column=col_idx, value=sh)
        cell.font = font_header
        cell.fill = fill_accent
        cell.alignment = align_center
        cell.border = border_thin
    
    ws_dash.row_dimensions[13].height = 24

    phases_summary = [
        ("Phase 1: Foundations of Embedded C", "Week 01 - 06", "=IF(COUNTIF('Weekly Progress Tracker'!D5:D10, \"Completed\")=6, \"Completed\", IF(COUNTIF('Weekly Progress Tracker'!D5:D10, \"Not Started\")=6, \"Not Started\", \"In Progress\"))", "=COUNTIF('Weekly Progress Tracker'!D5:D10, \"Completed\")/6", "Custom Allocator & Logic Gates Simulator"),
        ("Phase 2: Bare-Metal Driver Dev", "Week 07 - 12", "=IF(COUNTIF('Weekly Progress Tracker'!D11:D16, \"Completed\")=6, \"Completed\", IF(COUNTIF('Weekly Progress Tracker'!D11:D16, \"Not Started\")=6, \"Not Started\", \"In Progress\"))", "=COUNTIF('Weekly Progress Tracker'!D11:D16, \"Completed\")/6", "MPU6050 Accelerometer I2C Driver"),
        ("Phase 3: Linux Systems Programming", "Week 13 - 18", "=IF(COUNTIF('Weekly Progress Tracker'!D17:D22, \"Completed\")=6, \"Completed\", IF(COUNTIF('Weekly Progress Tracker'!D17:D22, \"Not Started\")=6, \"Not Started\", \"In Progress\"))", "=COUNTIF('Weekly Progress Tracker'!D17:D22, \"Completed\")/6", "Multi-Threaded IPC Packet Processor"),
        ("Phase 4: Real-Time OS (RTOS)", "Week 19 - 24", "=IF(COUNTIF('Weekly Progress Tracker'!D23:F28, \"Completed\")=6, \"Completed\", IF(COUNTIF('Weekly Progress Tracker'!D23:D28, \"Not Started\")=6, \"Not Started\", \"In Progress\"))", "=COUNTIF('Weekly Progress Tracker'!D23:D28, \"Completed\")/6", "RTOS Data Logger Watchdog Control"),
        ("Phase 5: Embedded Linux & Drivers", "Week 25 - 30", "=IF(COUNTIF('Weekly Progress Tracker'!D29:D34, \"Completed\")=6, \"Completed\", IF(COUNTIF('Weekly Progress Tracker'!D29:D34, \"Not Started\")=6, \"Not Started\", \"In Progress\"))", "=COUNTIF('Weekly Progress Tracker'!D29:D34, \"Completed\")/6", "LED & Button Interrupt Kernel Driver"),
        ("Phase 6: Yocto & Capstone Project", "Week 31 - 36", "=IF(COUNTIF('Weekly Progress Tracker'!D35:D40, \"Completed\")=6, \"Completed\", IF(COUNTIF('Weekly Progress Tracker'!D35:D40, \"Not Started\")=6, \"Not Started\", \"In Progress\"))", "=COUNTIF('Weekly Progress Tracker'!D35:D40, \"Completed\")/6", "IIoT FreeRTOS-Linux Gateway System")
    ]

    for idx, row_data in enumerate(phases_summary, 14):
        ws_dash.row_dimensions[idx].height = 20
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_dash.cell(row=idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = border_thin
            
            if col_idx in [1, 5]:
                cell.alignment = align_left
            elif col_idx in [2, 3]:
                cell.alignment = align_center
            elif col_idx == 4:
                cell.alignment = align_right
                cell.number_format = "0%"

            if col_idx == 3:
                cell.font = font_body_bold

        if idx % 2 == 1:
            for c in range(1, 6):
                ws_dash.cell(row=idx, column=c).fill = fill_zebra

    # Totals Row
    tot_row = 20
    ws_dash.cell(row=tot_row, column=1, value="Overall Training Program Status").font = font_body_bold
    ws_dash.cell(row=tot_row, column=1).alignment = align_left
    ws_dash.cell(row=tot_row, column=1).border = border_double_bottom
    
    ws_dash.cell(row=tot_row, column=2, value="36 Weeks").font = font_body_bold
    ws_dash.cell(row=tot_row, column=2).alignment = align_center
    ws_dash.cell(row=tot_row, column=2).border = border_double_bottom

    ws_dash.cell(row=tot_row, column=3, value="=IF(G10=1, \"Completed\", \"Active\")").font = font_body_bold
    ws_dash.cell(row=tot_row, column=3).alignment = align_center
    ws_dash.cell(row=tot_row, column=3).border = border_double_bottom
    
    ws_dash.cell(row=tot_row, column=4, value="=AVERAGE(D14:D19)").font = font_body_bold
    ws_dash.cell(row=tot_row, column=4).alignment = align_right
    ws_dash.cell(row=tot_row, column=4).number_format = "0%"
    ws_dash.cell(row=tot_row, column=4).border = border_double_bottom
    
    ws_dash.cell(row=tot_row, column=5, value="Final Portfolios Review").font = font_body_bold
    ws_dash.cell(row=tot_row, column=5).alignment = align_left
    ws_dash.cell(row=tot_row, column=5).border = border_double_bottom

    # Adjust widths for Dashboard
    ws_dash.column_dimensions["A"].width = 38
    ws_dash.column_dimensions["B"].width = 18
    ws_dash.column_dimensions["C"].width = 18
    ws_dash.column_dimensions["D"].width = 16
    ws_dash.column_dimensions["E"].width = 38
    ws_dash.column_dimensions["F"].width = 18
    ws_dash.column_dimensions["G"].width = 24

    # 5. Coach's Intervention and Actions Plan Area (Row 23 to 27)
    ws_dash.cell(row=22, column=1, value="COACH'S INTERVENTION PLAN & STRATEGIC ACTION LOG").font = font_section
    ws_dash.merge_cells("A23:G23")
    act_hdr = ws_dash.cell(row=23, column=1, value="Active Blockages & Mentor Directive Actions")
    act_hdr.font = font_header
    act_hdr.fill = fill_accent
    act_hdr.alignment = align_left
    act_hdr.border = border_thin

    ws_dash.merge_cells("A24:G25")
    act_body = ws_dash.cell(row=24, column=1, value="Sivaramireddy is performing exceptionally on Week 1 & 2. The next gating challenge is the static allocators mapping in Week 4. Directive: Ensure the student sketches the complete memory block maps before coding to prevent pointers misalignment bugs.")
    act_body.font = font_body
    act_body.alignment = align_left
    act_body.border = border_thin

    ws_dash.row_dimensions[22].height = 20
    ws_dash.row_dimensions[23].height = 22
    ws_dash.row_dimensions[24].height = 20
    ws_dash.row_dimensions[25].height = 20

    # -------------------------------------------------------------
    # Conditional Formattings
    # -------------------------------------------------------------
    # Status formatting
    fill_green = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    font_green = Font(name=FONT_FAMILY, size=10, bold=True, color="166534")
    
    fill_yellow = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
    font_yellow = Font(name=FONT_FAMILY, size=10, bold=True, color="854D0E")

    fill_gray = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    font_gray = Font(name=FONT_FAMILY, size=10, bold=True, color="475569")

    # Code quality formatting
    fill_fail = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    font_fail = Font(name=FONT_FAMILY, size=10, bold=True, color="991B1B")

    # Apply to Tasks Tracker Status Column (D5:D40)
    ws_tasks.conditional_formatting.add("D5:D40", CellIsRule(operator="equal", formula=['"Completed"'], fill=fill_green, font=font_green))
    ws_tasks.conditional_formatting.add("D5:D40", CellIsRule(operator="equal", formula=['"In Progress"'], fill=fill_yellow, font=font_yellow))
    ws_tasks.conditional_formatting.add("D5:D40", CellIsRule(operator="equal", formula=['"Not Started"'], fill=fill_gray, font=font_gray))

    # Apply to Dashboard Status Column (C14:C19)
    ws_dash.conditional_formatting.add("C14:C19", CellIsRule(operator="equal", formula=['"Completed"'], fill=fill_green, font=font_green))
    ws_dash.conditional_formatting.add("C14:C19", CellIsRule(operator="equal", formula=['"In Progress"'], fill=fill_yellow, font=font_yellow))
    ws_dash.conditional_formatting.add("C14:C19", CellIsRule(operator="equal", formula=['"Not Started"'], fill=fill_gray, font=font_gray))

    # Apply to Tasks Code Quality Column (F5:F40)
    ws_tasks.conditional_formatting.add("F5:F40", CellIsRule(operator="equal", formula=['"Pass"'], fill=fill_green, font=font_green))
    ws_tasks.conditional_formatting.add("F5:F40", CellIsRule(operator="equal", formula=['"Fail"'], fill=fill_fail, font=font_fail))
    ws_tasks.conditional_formatting.add("F5:F40", CellIsRule(operator="equal", formula=['"Pending"'], fill=fill_gray, font=font_gray))

    # Save
    out_path = "/home/siva/sivaramireddy/Project1/curriculum_progress_tracker.xlsx"
    wb.save(out_path)
    print(f"SUCCESS: Mentorship Progress Dashboard created at: {out_path}")

if __name__ == "__main__":
    main()
