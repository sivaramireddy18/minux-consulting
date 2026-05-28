import os
import sys

def main():
    # 1. Install openpyxl dynamically if missing
    try:
        import openpyxl
    except ImportError:
        print("Installing openpyxl library inside user context...")
        import subprocess
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--user", "--break-system-packages"])
        except Exception:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--user"])
        import site
        sys.path.append(site.getusersitepackages())
        import openpyxl

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule

    # Create Workbook
    wb = openpyxl.Workbook()
    
    # Setup Sheets
    ws_dash = wb.active
    ws_dash.title = "Executive Dashboard"
    ws_tasks = wb.create_sheet(title="Weekly Progress Tracker")

    # Ensure grid lines are visible
    ws_dash.views.sheetView[0].showGridLines = True
    ws_tasks.views.sheetView[0].showGridLines = True

    # -------------------------------------------------------------
    # Premium Colors & Styling Definitions (Teal & Slate Palette)
    # -------------------------------------------------------------
    FONT_FAMILY = "Segoe UI"
    
    # Fonts
    font_title = Font(name=FONT_FAMILY, size=18, bold=True, color="0F172A")
    font_section = Font(name=FONT_FAMILY, size=13, bold=True, color="0F766E")
    font_header = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")
    font_body = Font(name=FONT_FAMILY, size=10, color="334155")
    font_body_bold = Font(name=FONT_FAMILY, size=10, bold=True, color="1E293B")
    font_kpi_num = Font(name=FONT_FAMILY, size=20, bold=True, color="0F766E")
    font_kpi_lbl = Font(name=FONT_FAMILY, size=9, bold=True, color="64748B")

    # Fills
    fill_primary = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid") # Dark Teal
    fill_accent = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Slate
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Light Slate
    fill_kpi = PatternFill(start_color="F0FDFA", end_color="F0FDFA", fill_type="solid") # Very Light Teal
    fill_sub_hdr = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid") # Light Gray

    # Alignments
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")
    
    # Borders
    thin_border_side = Side(border_style="thin", color="CBD5E1")
    border_thin = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_kpi = Border(left=Side(style="medium", color="0F766E"), right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    double_bottom = Border(bottom=Side(style="double", color="0F172A"), top=Side(style="thin", color="CBD5E1"))

    # -------------------------------------------------------------
    # Populate Sheet 2: Weekly Progress Tracker
    # -------------------------------------------------------------
    headers = [
        "Phase", "Week", "Core Focus Area", "Milestone Gating Challenge", 
        "Target Date", "Status", "Weight", "Grade / Code Quality", "Comments / Blockages"
    ]
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws_tasks.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = align_center
        cell.border = border_thin
    
    ws_tasks.row_dimensions[4].height = 28

    # Curriculum Data Definition
    curriculum = [
        ("Phase 1: Foundations", "Week 01", "Git Workflows, Repo Discipline, and Dev Setup", "Custom pre-commit shell validation hook", "2026-06-05", 0.05),
        ("Phase 1: Foundations", "Week 02", "Digital Logic, Radix Formats, Boolean Gating", "4-Bit Binary Full-Adder Emulator", "2026-06-12", 0.05),
        ("Phase 1: Foundations", "Week 03", "Embedded C Basics, Variable Lifetimes, Volatile", "Incremental Make build system", "2026-06-19", 0.05),
        ("Phase 1: Foundations", "Week 04", "Pointers, App Memory Maps, Memory Allocators", "Raw Static Array Memory Allocator", "2026-06-26", 0.05),
        ("Phase 1: Foundations", "Week 05", "Bitwise Math, Bitfields, Registers MMIO", "GPIO Register Controller Emulator", "2026-07-03", 0.05),
        ("Phase 1: Foundations", "Week 06", "Structures Packing, Alignment, Serialization", "Serialized Frame Packetizer (CRC-8)", "2026-07-10", 0.05),
        
        ("Phase 2: Bare-Metal", "Week 07", "ARM Cortex-M Core, Linker Scripts, Vectors", "Reset Handler Startup File from scratch", "2026-07-17", 0.05),
        ("Phase 2: Bare-Metal", "Week 08", "Bare-Metal Clocks Gating & GPIO Registers", "Register-Level Button LED Debouncer", "2026-07-24", 0.05),
        ("Phase 2: Bare-Metal", "Week 09", "Interrupts Controller (NVIC), EXTI pin routing", "Critical Section Protected ISR Handler", "2026-07-31", 0.05),
        ("Phase 2: Bare-Metal", "Week 10", "SysTick Counter, General TIM interrupts, ARR/PSC", "TIM2 periodic timer (500ms heartbeat)", "2026-08-07", 0.05),
        ("Phase 2: Bare-Metal", "Week 11", "Communication Protocols I: UART Serial Drivers", "Interrupt-driven RX Ring Buffer (115200)", "2026-08-14", 0.05),
        ("Phase 2: Bare-Metal", "Week 12", "Communication Protocols II: I2C & SPI Registers", "Bare-Metal MPU6050 Accelerometer Driver", "2026-08-21", 0.05),
        
        ("Phase 3: Systems Prog", "Week 13", "Linux OS CLI Command Pipelines, FHS, Perms", "System Diagnostics Dashboard Script", "2026-08-28", 0.05),
        ("Phase 3: Systems Prog", "Week 14", "GCC compiler stages, Makefile dependencies, ELF", "Symbol Dissection Shell Script (readelf/nm)", "2026-09-04", 0.05),
        ("Phase 3: Systems Prog", "Week 15", "Low-level system I/O, ioctl, memory-mapping", "Memory-Mapped File Database Processor", "2026-09-11", 0.05),
        ("Phase 3: Systems Prog", "Week 16", "Process Control, fork, exec, wait, Shared Memory", "Multi-Process Real-time IPC Gateway", "2026-09-18", 0.05),
        ("Phase 3: Systems Prog", "Week 17", "POSIX threads, mutex, condition variables", "Thread-Safe Concurrent FIFO Queue", "2026-09-25", 0.05),
        ("Phase 3: Systems Prog", "Week 18", "Bash Automation scripts, sed stream, awk parsing", "CI/CD Build, Audit & Test Shell Suite", "2026-10-02", 0.05),
        
        ("Phase 4: RTOS Core", "Week 19", "RTOS Paradigm, FreeRTOS ports, TCB structure", "Multi-tasking concurrent blinker port", "2026-10-09", 0.05),
        ("Phase 4: RTOS Core", "Week 20", "Task state machine, preemptive scheduler, PendSV", "Assembly Context Switch simulator", "2026-10-16", 0.05),
        ("Phase 4: RTOS Core", "Week 21", "Binary/Counting Semaphores, Mutexes, Inversion", "Priority Inheritance Mutex protect driver", "2026-10-23", 0.05),
        ("Phase 4: RTOS Core", "Week 22", "ITC, Message Queues by value, Event Groups", "Multi-Sensor Telemetry Concentrator", "2026-10-30", 0.05),
        ("Phase 4: RTOS Core", "Week 23", "Memory schemes (heap_1 to 5), stack overflow", "RTOS Memory Sentinel Stack check hook", "2026-11-06", 0.05),
        ("Phase 4: RTOS Core", "Week 24", "Software timers, ISR FromISR APIs, deferred", "Deferred Interrupt Button Debouncer", "2026-11-13", 0.05),
        
        ("Phase 5: Kernel Space", "Week 25", "Embedded Linux bootloader, SPL, U-Boot CLI", "Boot Log Diagnostic parser script", "2026-11-20", 0.05),
        ("Phase 5: Kernel Space", "Week 26", "Cross-compilation, Kernel source compile, boot", "Custom Kernel Image (zImage) compilation", "2026-11-27", 0.05),
        ("Phase 5: Kernel Space", "Week 27", "Device Tree DTS/DTSI node maps, compatible", "Dynamic Device Tree Overlay (.dtbo) compile", "2026-12-04", 0.05),
        ("Phase 5: Kernel Space", "Week 28", "Out-of-tree loadable module (LKM), parameters", "Dynamic Parameter Telemetry Kernel Driver", "2026-12-11", 0.05),
        ("Phase 5: Kernel Space", "Week 29", "Character Device registration, cdev, file_ops", "Virtual Ring Buffer Character Driver", "2026-12-18", 0.05),
        ("Phase 5: Kernel Space", "Week 30", "remap virtual kernel map, legacy GPIO, IRQs", "LED and Button Interrupt Kernel Driver", "2026-12-25", 0.05),
        
        ("Phase 6: Integration", "Week 31", "Rootfs generation Buildroot, overlays, packages", "Minimal custom Embedded Linux Image", "2027-01-01", 0.05),
        ("Phase 6: Integration", "Week 32", "Yocto Poky layers, BitBake compile, recipes", "Custom Gateway layer & recipe config", "2027-01-08", 0.05),
        ("Phase 6: Integration", "Week 33", "Analog-to-digital (ADC), register PWM control", "Closed-loop ADC-PWM Dimmer controller", "2027-01-15", 0.05),
        ("Phase 6: Integration", "Week 34", "POSIX Sockets, TCP/IP concurrent multi-thread", "Multi-Client TCP Telemetry Server in C", "2027-01-22", 0.05),
        ("Phase 6: Integration", "Week 35", "System profiling, gdbserver, valgrind, perf", "System Diagnostics Audit and optimization", "2027-01-29", 0.05),
        ("Phase 6: Integration", "Week 36", "Capstone IoT Gateway integration, Yocto boot", "Industrial IoT Gateway system (FreeRTOS-Linux)", "2027-02-05", 0.05),
    ]

    for idx, row_data in enumerate(curriculum, 5):
        ws_tasks.row_dimensions[idx].height = 20
        # Write values
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_tasks.cell(row=idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = border_thin
            
            # Alignments & formats
            if col_idx in [1, 3, 4]:
                cell.alignment = align_left
            elif col_idx == 2:
                cell.alignment = align_center
                cell.font = font_body_bold
            elif col_idx == 5:
                cell.alignment = align_center
            elif col_idx == 7:
                cell.alignment = align_right
                cell.number_format = "0%"
        
        # Default empty tracking values
        status_cell = ws_tasks.cell(row=idx, column=6, value="Not Started")
        status_cell.font = font_body_bold
        status_cell.alignment = align_center
        status_cell.border = border_thin
        
        grade_cell = ws_tasks.cell(row=idx, column=8, value="Pending")
        grade_cell.font = font_body
        grade_cell.alignment = align_center
        grade_cell.border = border_thin
        
        comm_cell = ws_tasks.cell(row=idx, column=9, value="")
        comm_cell.border = border_thin
        
        # Apply zebra striping
        if idx % 2 == 0:
            for col_idx in range(1, 10):
                ws_tasks.cell(row=idx, column=col_idx).fill = fill_zebra

    # Add drop-down list validation to Status column
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"Not Started,In Progress,Completed"', allow_blank=True)
    ws_tasks.add_data_validation(dv)
    dv.add(f"F5:F40")

    # Add title header in Tasks list
    ws_tasks.merge_cells("A1:I2")
    title_cell = ws_tasks.cell(row=1, column=1, value="EMBEDDED SYSTEMS & LINUX CURRICULUM - WEEKLY TRACKER")
    title_cell.font = font_title
    title_cell.alignment = align_center

    # Auto-adjust column widths for Weekly Tasks
    for col in ws_tasks.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Skip title
            if cell.row in [1, 2]: continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws_tasks.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # -------------------------------------------------------------
    # Populate Sheet 1: Executive Dashboard
    # -------------------------------------------------------------
    # Header block
    ws_dash.merge_cells("A1:G2")
    dash_title = ws_dash.cell(row=1, column=1, value="CURRICULUM TRAINING PROGRESS DASHBOARD")
    dash_title.font = font_title
    dash_title.alignment = align_center
    
    ws_dash.row_dimensions[1].height = 20
    ws_dash.row_dimensions[2].height = 20

    # 1. KPI Blocks Header (Merge-styled card blocks)
    kpis = [
        ("TOTAL WEEKS", "36", "A4:B5"),
        ("COMPLETED WEEKS", "=COUNTIF('Weekly Progress Tracker'!F5:F40, \"Completed\")", "C4:D5"),
        ("IN PROGRESS", "=COUNTIF('Weekly Progress Tracker'!F5:F40, \"In Progress\")", "E4:F5"),
        ("COMPLETION RATE", "='Weekly Progress Tracker'!B37", "G4:G5") # Dynamic completion rate link (or formula)
    ]
    
    # We will write the formulas in a clean structure:
    # Row 4: KPI Label
    # Row 5: KPI Value (Formula linked)
    
    def setup_kpi_card(ws, label, formula, col_start, col_end):
        ws.merge_cells(f"{col_start}4:{col_end}4")
        lbl_cell = ws[f"{col_start}4"]
        lbl_cell.value = label
        lbl_cell.font = font_kpi_lbl
        lbl_cell.alignment = align_center
        lbl_cell.fill = fill_kpi
        
        ws.merge_cells(f"{col_start}5:{col_end}5")
        val_cell = ws[f"{col_start}5"]
        val_cell.value = formula
        val_cell.font = font_kpi_num
        val_cell.alignment = align_center
        val_cell.fill = fill_kpi
        
        # Apply borders to card area
        for r in [4, 5]:
            for c in range(openpyxl.utils.column_index_from_string(col_start), openpyxl.utils.column_index_from_string(col_end) + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border_kpi

    setup_kpi_card(ws_dash, "TOTAL WEEKS", 36, "A", "B")
    setup_kpi_card(ws_dash, "COMPLETED", "=COUNTIF('Weekly Progress Tracker'!F5:F40, \"Completed\")", "C", "C")
    setup_kpi_card(ws_dash, "IN PROGRESS", "=COUNTIF('Weekly Progress Tracker'!F5:F40, \"In Progress\")", "D", "D")
    setup_kpi_card(ws_dash, "NOT STARTED", "=COUNTIF('Weekly Progress Tracker'!F5:F40, \"Not Started\")", "E", "E")
    
    # Total completion percentage card
    ws_dash.merge_cells("F4:G4")
    lbl_comp = ws_dash.cell(row=4, column=6, value="OVERALL COMPLETION %")
    lbl_comp.font = font_kpi_lbl
    lbl_comp.alignment = align_center
    lbl_comp.fill = fill_kpi
    lbl_comp.border = border_kpi
    
    ws_dash.merge_cells("F5:G5")
    val_comp = ws_dash.cell(row=5, column=6, value="=C5/A5")
    val_comp.font = font_kpi_num
    val_comp.alignment = align_center
    val_comp.fill = fill_kpi
    val_comp.number_format = "0.0%"
    val_comp.border = border_kpi

    ws_dash.row_dimensions[4].height = 18
    ws_dash.row_dimensions[5].height = 32

    # 2. Phase-wise Summary Table
    ws_dash.cell(row=8, column=1, value="TRAINING PHASE PERFORMANCE HIGHLIGHTS").font = font_section
    
    summary_headers = ["Phase / Module Name", "Weeks Included", "Status", "Completion", "Milestone Target"]
    for col_idx, sh in enumerate(summary_headers, 1):
        cell = ws_dash.cell(row=9, column=col_idx, value=sh)
        cell.font = font_header
        cell.fill = fill_accent
        cell.alignment = align_center
        cell.border = border_thin
    
    ws_dash.row_dimensions[9].height = 24

    phases_summary = [
        ("Phase 1: Foundations of Embedded C", "Week 01 - 06", "=IF(COUNTIF('Weekly Progress Tracker'!F5:F10, \"Completed\")=6, \"Completed\", IF(COUNTIF('Weekly Progress Tracker'!F5:F10, \"Not Started\")=6, \"Not Started\", \"In Progress\"))", "=COUNTIF('Weekly Progress Tracker'!F5:F10, \"Completed\")/6", "Custom Allocator & Logic Gates Simulator"),
        ("Phase 2: Bare-Metal Driver Dev", "Week 07 - 12", "=IF(COUNTIF('Weekly Progress Tracker'!F11:F16, \"Completed\")=6, \"Completed\", IF(COUNTIF('Weekly Progress Tracker'!F11:F16, \"Not Started\")=6, \"Not Started\", \"In Progress\"))", "=COUNTIF('Weekly Progress Tracker'!F11:F16, \"Completed\")/6", "MPU6050 Accelerometer I2C Driver"),
        ("Phase 3: Linux Systems Programming", "Week 13 - 18", "=IF(COUNTIF('Weekly Progress Tracker'!F17:F22, \"Completed\")=6, \"Completed\", IF(COUNTIF('Weekly Progress Tracker'!F17:F22, \"Not Started\")=6, \"Not Started\", \"In Progress\"))", "=COUNTIF('Weekly Progress Tracker'!F17:F22, \"Completed\")/6", "Multi-Threaded IPC Packet Processor"),
        ("Phase 4: Real-Time OS (RTOS)", "Week 19 - 24", "=IF(COUNTIF('Weekly Progress Tracker'!F23:F28, \"Completed\")=6, \"Completed\", IF(COUNTIF('Weekly Progress Tracker'!F23:F28, \"Not Started\")=6, \"Not Started\", \"In Progress\"))", "=COUNTIF('Weekly Progress Tracker'!F23:F28, \"Completed\")/6", "RTOS Data Logger Watchdog Control"),
        ("Phase 5: Embedded Linux & Drivers", "Week 25 - 30", "=IF(COUNTIF('Weekly Progress Tracker'!F29:F34, \"Completed\")=6, \"Completed\", IF(COUNTIF('Weekly Progress Tracker'!F29:F34, \"Not Started\")=6, \"Not Started\", \"In Progress\"))", "=COUNTIF('Weekly Progress Tracker'!F29:F34, \"Completed\")/6", "LED & Button Interrupt Kernel Driver"),
        ("Phase 6: Yocto & Capstone Project", "Week 31 - 36", "=IF(COUNTIF('Weekly Progress Tracker'!F35:F40, \"Completed\")=6, \"Completed\", IF(COUNTIF('Weekly Progress Tracker'!F35:F40, \"Not Started\")=6, \"Not Started\", \"In Progress\"))", "=COUNTIF('Weekly Progress Tracker'!F35:F40, \"Completed\")/6", "IIoT FreeRTOS-Linux Gateway System")
    ]

    for idx, row_data in enumerate(phases_summary, 10):
        ws_dash.row_dimensions[idx].height = 22
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_dash.cell(row=idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = border_thin
            
            if col_idx in [1, 5]:
                cell.alignment = align_left
            elif col_idx in [2, 3]:
                cell.alignment = align_center
            elif col_idx == 4:
                cell.alignment = align_right
                cell.number_format = "0%"

            if col_idx == 3:
                cell.font = font_body_bold

        if idx % 2 == 1:
            for c in range(1, 6):
                ws_dash.cell(row=idx, column=c).fill = fill_zebra

    # Add totals row to Dashboard
    tot_row = 16
    ws_dash.cell(row=tot_row, column=1, value="Overall Program Status").font = font_body_bold
    ws_dash.cell(row=tot_row, column=1).alignment = align_left
    ws_dash.cell(row=tot_row, column=1).border = double_bottom
    
    ws_dash.cell(row=tot_row, column=2, value="36 Weeks").font = font_body_bold
    ws_dash.cell(row=tot_row, column=2).alignment = align_center
    ws_dash.cell(row=tot_row, column=2).border = double_bottom

    ws_dash.cell(row=tot_row, column=3, value="=IF(F5=1, \"Completed\", \"Active\")").font = font_body_bold
    ws_dash.cell(row=tot_row, column=3).alignment = align_center
    ws_dash.cell(row=tot_row, column=3).border = double_bottom
    
    ws_dash.cell(row=tot_row, column=4, value="=AVERAGE(D10:D15)").font = font_body_bold
    ws_dash.cell(row=tot_row, column=4).alignment = align_right
    ws_dash.cell(row=tot_row, column=4).number_format = "0%"
    ws_dash.cell(row=tot_row, column=4).border = double_bottom
    
    ws_dash.cell(row=tot_row, column=5, value="Final Graduation").font = font_body_bold
    ws_dash.cell(row=tot_row, column=5).alignment = align_left
    ws_dash.cell(row=tot_row, column=5).border = double_bottom

    # Adjust widths
    ws_dash.column_dimensions["A"].width = 38
    ws_dash.column_dimensions["B"].width = 18
    ws_dash.column_dimensions["C"].width = 18
    ws_dash.column_dimensions["D"].width = 16
    ws_dash.column_dimensions["E"].width = 38
    ws_dash.column_dimensions["F"].width = 14
    ws_dash.column_dimensions["G"].width = 14

    # -------------------------------------------------------------
    # Conditional Formatting (Completed vs In Progress colors)
    # -------------------------------------------------------------
    # Fills
    fill_green = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # completed
    font_green = Font(name=FONT_FAMILY, size=10, bold=True, color="166534")
    
    fill_yellow = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid") # in progress
    font_yellow = Font(name=FONT_FAMILY, size=10, bold=True, color="854D0E")

    fill_gray = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid") # not started
    font_gray = Font(name=FONT_FAMILY, size=10, bold=True, color="475569")

    # Apply to Tasks Tracker Status Column (F5:F40)
    ws_tasks.conditional_formatting.add("F5:F40", CellIsRule(operator="equal", formula=['"Completed"'], fill=fill_green, font=font_green))
    ws_tasks.conditional_formatting.add("F5:F40", CellIsRule(operator="equal", formula=['"In Progress"'], fill=fill_yellow, font=font_yellow))
    ws_tasks.conditional_formatting.add("F5:F40", CellIsRule(operator="equal", formula=['"Not Started"'], fill=fill_gray, font=font_gray))

    # Apply to Dashboard Status Column (C10:C15)
    ws_dash.conditional_formatting.add("C10:C15", CellIsRule(operator="equal", formula=['"Completed"'], fill=fill_green, font=font_green))
    ws_dash.conditional_formatting.add("C10:C15", CellIsRule(operator="equal", formula=['"In Progress"'], fill=fill_yellow, font=font_yellow))
    ws_dash.conditional_formatting.add("C10:C15", CellIsRule(operator="equal", formula=['"Not Started"'], fill=fill_gray, font=font_gray))

    # Save Workbook
    out_path = "/home/siva/sivaramireddy/Project1/curriculum_progress_tracker.xlsx"
    wb.save(out_path)
    print(f"SUCCESS: Highly styled, professional Excel Tracker generated at: {out_path}")

if __name__ == "__main__":
    main()
