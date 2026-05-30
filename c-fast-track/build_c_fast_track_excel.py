#!/usr/bin/env python3
import os
import sys
import datetime

def main():
    WORKSPACE = "/home/siva/sivaramireddy/Project1"
    C_FAST_TRACK_DIR = os.path.join(WORKSPACE, "c-fast-track")
    # 1. Install openpyxl dynamically if missing
    try:
        import openpyxl
    except ImportError:
        print("Installing openpyxl library inside user context...")
        import subprocess
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--user", "--break-system-packages"])
        except Exception:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--user"])
        import site
        sys.path.append(site.getusersitepackages())
        import openpyxl

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.fill import PatternFillProperties, ColorChoice

    # Create Workbook
    wb = openpyxl.Workbook()
    
    # Setup Sheets
    ws_intro = wb.active
    ws_intro.title = "User Guide & Vetting Guide"
    ws_dash = wb.create_sheet(title="Executive Dashboard")
    ws_tasks = wb.create_sheet(title="Daily Progress Tracker")

    # Ensure grid lines are visible
    ws_intro.views.sheetView[0].showGridLines = True
    ws_dash.views.sheetView[0].showGridLines = True
    ws_tasks.views.sheetView[0].showGridLines = True

    # -------------------------------------------------------------
    # Premium Colors & Styling Definitions (Teal & Slate Palette)
    # -------------------------------------------------------------
    FONT_FAMILY = "Segoe UI"
    
    # Fonts
    font_title = Font(name=FONT_FAMILY, size=18, bold=True, color="0F172A")
    font_section = Font(name=FONT_FAMILY, size=13, bold=True, color="0F766E")
    font_header = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")
    font_body = Font(name=FONT_FAMILY, size=10, color="334155")
    font_body_bold = Font(name=FONT_FAMILY, size=10, bold=True, color="1E293B")
    font_kpi_num = Font(name=FONT_FAMILY, size=20, bold=True, color="0F766E")
    font_kpi_lbl = Font(name=FONT_FAMILY, size=9, bold=True, color="64748B")

    # Fills
    fill_primary = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid") # Dark Teal
    fill_accent = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Slate
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Light Slate
    fill_kpi = PatternFill(start_color="F0FDFA", end_color="F0FDFA", fill_type="solid") # Very Light Teal
    fill_sub_hdr = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid") # Light Gray

    # Alignments
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")
    
    # Borders
    thin_border_side = Side(border_style="thin", color="CBD5E1")
    border_thin = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_kpi = Border(left=Side(style="medium", color="0F766E"), right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    double_bottom = Border(bottom=Side(style="double", color="0F172A"), top=Side(style="thin", color="CBD5E1"))

    # -------------------------------------------------------------
    # 45-Day Curriculum Data Compilation
    # -------------------------------------------------------------
    phases = [
        ("Phase 1: Compiler & Toolchains", [
            ("The Preprocessor Pipeline and Macro Expansion Hazards", "Compile and audit square preprocessor macros side-effects output"),
            ("Compilation to Assembly and GDB Register Analysis", "Compile to assembly using -S and step through execution registers"),
            ("Relocatable Object Files Dissection and Symbol Resolution", "Dissect symbols and segments (.text, .data, .bss, .rodata) using nm"),
            ("Linking Mechanics, Static/Dynamic Libraries, and Symbol Collisions", "Compile and link dynamic libraries checking rpath"),
            ("GNU Make, Implicit Dependencies, and Incremental Builds", "Write self-dependency tracking Makefile with MM/MMD")
        ]),
        ("Phase 2: Data & Radix Hazards", [
            ("POSIX Integer Scales, Two's Complement, and Sign Extensions", "Assign signed variables to unsigned variables printing raw hexadecimal"),
            ("Safe Integer Arithmetic, Overflow Detection, and Type Promotions", "Write robust overflow pre-check addition/multiplication C functions"),
            ("Floating-Point Internals (IEEE 754) and Precision Hazards", "Write robust epsilon float equality comparison logic"),
            ("Bitwise Operations, Masking, and Shift Hazards", "Design register status flags bit togglers avoiding undefined shifts"),
            ("Branch-Free Programming and ALU Pipeline Optimization", "Write branch-free absolute value and select methods checking assembly")
        ]),
        ("Phase 3: Stack & Execution Frame", [
            ("Activation Records and Procedure Call Standards (AAPCS)", "Verify AAPCS register arguments passing in GDB breakpoints"),
            ("Function Frame Layout, Prologue, and Epilogue Assembly", "Audit SP moves and push/pop instruction layouts inside custom assembly"),
            ("Local Variable Lifetimes, Storage Classes, and Scope Boundaries", "Verify static address persistence vs stack decay warnings"),
            ("Stack Overflow Hazards, MSP vs PSP, and MPU Guard Bands", "Generate stack overflows in GDB using recursive allocations"),
            ("Stack Canaries and Buffer Overflow Defensive Coding", "Verify stack protector failure trap hook by smashing local buffer")
        ]),
        ("Phase 4: Pointers & Arithmetic", [
            ("Pointers Dereferencing, Decay, and Multi-Dimensional Arrays", "Write row-major calculations representing matrix offsets"),
            ("Pointer Arithmetic Scaling Mathematics and Offset Mapping", "Verify address scaling gaps based on target types size differences"),
            ("Volatile Pointer Configurations and Memory-Mapped Registers", "Audit optimized assembly polling loops with and without volatile"),
            ("Function Pointers, Callback Systems, and Jump-Tables", "Implement calculator command jump-table eliminating switch logic"),
            ("Raw Pointer Casting, Type Safety, and Alignment Validations", "Design address checks tracking alignment gaps before pointer writes")
        ]),
        ("Phase 5: Structures & Alignments", [
            ("Structure Memory Layouts, Compiler Padding, and Offset Tracing", "Trace struct layout offsets to minimize padding size footprint"),
            ("Packed Structures, Alignment Pragma/Attributes, and Performance Costs", "Analyze normal vs packed structures and assembly operations cost"),
            ("Unions, Type Punning, and Raw Binary Serialization", "Inspect raw floating-point bits and serialize structs in unions"),
            ("Bitfields, Hardware Register Mapping, and Alignments", "Define status register bitfields and audit compile mask calculations"),
            ("Network Serialization, Endianness, and Byte Order Conversion", "Implement byte order serializer converting variables to Big-Endian")
        ]),
        ("Phase 6: Custom Allocators", [
            ("Dynamic Heap Architecture, Fragmentation, and System Boundaries", "Verify heap break address adjustments using sbrk system calls"),
            ("Custom Memory Arenas and Alignment Boundaries", "Implement linear aligned arena allocator enforcing 8-byte boundaries"),
            ("Fixed-Size Memory Block Pools", "Implement deterministic O(1) free list block pool memory slab allocator"),
            ("Dynamic Free Lists, Block Splitting, and Coalescing", "Design first-fit block coalescing merging neighboring free headers"),
            ("Memory Leaks Audits and Dynamic Wrapper Sentinels", "Implement custom malloc/free wrapper tracking allocations leaks context")
        ]),
        ("Phase 7: POSIX System Boundaries", [
            ("System Call Boundaries and Unbuffered POSIX I/O", "Write unbuffered file copier using open/read/write system calls"),
            ("Asynchronous I/O, Non-Blocking Descriptors, and Streaming", "Configure pipe descriptors in O_NONBLOCK mode handling EAGAIN"),
            ("Memory-Mapped Files (mmap) and Zero-Copy I/O", "Write database binary indexing editor using shared mmap maps"),
            ("System Error Handling, Thread-Safe errno, and Recovery", "Implement thread-safe strerror_r system error logger"),
            ("Inter-Process Communication (IPC) and Shared Memory", "Implement telemetry sender and receiver using POSIX shared memory")
        ]),
        ("Phase 8: Preprocessor Metaprogramming", [
            ("Preprocessor Extensions, Token Concatenation, and Stringification", "Write dynamic variable builder and string logger macros"),
            ("Multi-Line Macros Safety and do-while wrappers", "Wrap macro operations in do-while structures to pass conditional parses"),
            ("X-Macros Code Generation and Dynamic Lookups", "Generate enum declarations and lookup strings from central X-Macro list"),
            ("Variadic Macros and C11 _Generic Type Selections", "Design generic abs macros routing floats/ints at compile-time"),
            ("Include Guards, Conditional Compilations, and Assertions", "Enforce compile-time data sizes constraints using C11 _Static_assert")
        ]),
        ("Phase 9: Concurrency & MISRA C", [
            ("POSIX Threads (pthreads) Creation, Joins, and Configurations", "Spawn 4 worker pthreads executing concurrent jobs passing exits"),
            ("Race Conditions, Mutex Locks, and Deadlock Avoidance", "Audit counter racing under Helgrind and lock access using mutexes"),
            ("Condition Variables and Producer-Consumer Buffers", "Coordinate Producer-Consumer ring buffer utilizing pthreads cond variables"),
            ("MISRA C:2012 Guidelines and Critical Safety Constraints", "Refactor raw pointer operations to conform to safe MISRA rules"),
            ("Systems Profiling, Leak Audits, and Graduate Code Review", "Trace execution performance using gprof and debug leaks with Valgrind")
        ])
    ]

    # Generate dates (skipping weekends starting from 2026-06-01)
    start_date = datetime.date(2026, 6, 1)
    current_date = start_date
    dates = []
    while len(dates) < 45:
        if current_date.weekday() < 5: # Monday to Friday
            dates.append(current_date.strftime("%Y-%m-%d"))
        current_date += datetime.timedelta(days=1)

    # -------------------------------------------------------------
    # Populate Sheet 1: User Guide & Vetting Guide (Tab 1)
    # -------------------------------------------------------------
    ws_intro.merge_cells("A1:C2")
    intro_title = ws_intro.cell(row=1, column=1, value="C SYSTEMS PROGRAMMING 45-DAY ELITE FAST-TRACK - GUIDE & METRICS")
    intro_title.font = font_title
    intro_title.alignment = align_center
    intro_title.fill = fill_primary
    
    # Apply styling to merged title cells
    for r in range(1, 3):
        for c in range(1, 4):
            ws_intro.cell(row=r, column=c).fill = fill_primary
            ws_intro.cell(row=r, column=c).font = font_header

    ws_intro.row_dimensions[1].height = 20
    ws_intro.row_dimensions[2].height = 20

    # Section 1: Navigation Guide
    ws_intro.cell(row=4, column=1, value="1. HOW TO NAVIGATE AND FILL THE TRACKER").font = font_section
    ws_intro.row_dimensions[4].height = 24
    
    nav_headers = ["Action Step", "Procedural Instructions & Methodology", "Associated Tab / Location"]
    for col_idx, nh in enumerate(nav_headers, 1):
        cell = ws_intro.cell(row=5, column=col_idx, value=nh)
        cell.font = font_header
        cell.fill = fill_accent
        cell.alignment = align_center
        cell.border = border_thin
    ws_intro.row_dimensions[5].height = 22

    nav_steps = [
        ("Step 1: Read Daily Spec", "Go to your workspace folder: /c-fast-track/Day_Plans/. Read the corresponding daily blueprint (Day_01.md to Day_45.md). Read the 4-Hour Theory Deep-Dive and complete the 4-Hour Unassisted Lab.", "Workspace Files: Day_Plans/"),
        ("Step 2: Update Progress Status", "Select the 'Daily Progress Tracker' worksheet tab in this workbook. Locate the current day and set the 'Status' column dropdown selector to 'In Progress' or 'Completed'.", "Tab 3: Daily Progress Tracker"),
        ("Step 3: Audit & Grade Task", "Once the lab compiles and executes with zero warning flags, evaluate your understanding against the 5-Tier Vetting Scale. Select the appropriate rating from the 'Grade / Audit' column dropdown list.", "Tab 3: Daily Progress Tracker"),
        ("Step 4: Analyze Dashboard", "Switch to the 'Executive Dashboard' worksheet tab. Analyze real-time KPI metrics cards and track active phase completion progress displayed on the column charts.", "Tab 2: Executive Dashboard")
    ]

    for idx, row_data in enumerate(nav_steps, 6):
        ws_intro.row_dimensions[idx].height = 22
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_intro.cell(row=idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = border_thin
            if col_idx in [1, 2]:
                cell.alignment = align_left
            else:
                cell.alignment = align_center
        if idx % 2 == 1:
            for c in range(1, 4):
                ws_intro.cell(row=idx, column=c).fill = fill_zebra

    # Section 2: Vetting Criteria (Understanding Measurement)
    ws_intro.cell(row=12, column=1, value="2. 5-TIER UNDERSTANDING MEASUREMENT METRICS (VETTING CRITERIA)").font = font_section
    ws_intro.row_dimensions[12].height = 24

    v_headers = ["Grade / Audit Score", "Scale Target %", "Physical Evaluation Benchmarks (MISRA C & Hardware Vetting Standards)"]
    for col_idx, vh in enumerate(v_headers, 1):
        cell = ws_intro.cell(row=13, column=col_idx, value=vh)
        cell.font = font_header
        cell.fill = fill_accent
        cell.alignment = align_center
        cell.border = border_thin
    ws_intro.row_dimensions[13].height = 22

    vetting_criteria = [
        ("A+ [100%]", "100%", "Perfect unassisted delivery. Zero compiler warnings under strict '-Wall -Wextra -Werror -pedantic' C11 constraints. No memory leaks detected under Valgrind. Registers and hardware metrics successfully audited via GDB/Logic Analyzer."),
        ("A [90%]", "90%", "Clean compilation and execution with zero compiler warnings. Minor assisted code style adjustments permitted. High-quality documentation and safe defensive pointer boundaries implemented successfully."),
        ("B [80%]", "80%", "Compiles cleanly. Minor assisted corrections needed during structural alignments or preprocessor macros expansions. Program achieves functional output targets under debug audits."),
        ("C [70%]", "70%", "Compiles and runs with helper hints during lab execution. Program is functional, but contains suboptimal branching choices, redundant loops, or unoptimized variables structures."),
        ("F [Fail]", "0%", "Program fails to compile cleanly, contains memory leaks under Valgrind, or fails to implement basic functional registers validation gates."),
        ("Pending", "-", "Day task is not yet started or has not been audited.")
    ]

    for idx, row_data in enumerate(vetting_criteria, 14):
        ws_intro.row_dimensions[idx].height = 24
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_intro.cell(row=idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = border_thin
            if col_idx == 1:
                cell.font = font_body_bold
                cell.alignment = align_center
            elif col_idx == 2:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
        if idx % 2 == 1:
            for c in range(1, 4):
                ws_intro.cell(row=idx, column=c).fill = fill_zebra

    # Section 3: Summary Gating Checklist
    ws_intro.cell(row=22, column=1, value="3. KEY MILESTONE GATING CHECKS").font = font_section
    ws_intro.row_dimensions[22].height = 24

    gate_headers = ["Phase", "Days Mapped", "Core Gating Challenge Target"]
    for col_idx, gh in enumerate(gate_headers, 1):
        cell = ws_intro.cell(row=23, column=col_idx, value=gh)
        cell.font = font_header
        cell.fill = fill_accent
        cell.alignment = align_center
        cell.border = border_thin
    ws_intro.row_dimensions[23].height = 22

    gate_checklists = [
        ("Phase 1: Compiler & Toolchains", "Days 01 - 05", "Build modular project compiled with MM/MMD incremental Make dependency tracking."),
        ("Phase 2: Data & Radix Hazards", "Days 06 - 10", "Write branch-free absolute value/conditional select routines checking asm loops."),
        ("Phase 3: Stack & Execution Frame", "Days 11 - 15", "SMASH stack buffer of a dummy string array verifying stack canaries trap hook."),
        ("Phase 4: Pointers & Arithmetic", "Days 16 - 20", "Implement command processor jump-table routing calculations cleanly without switch constructs."),
        ("Phase 5: Structures & Alignments", "Days 21 - 25", "Manually serialize structural elements into Big-Endian arrays for network transport."),
        ("Phase 6: Custom Allocators", "Days 26 - 30", "Write custom aligned linear Arena and O(1) Slab allocator free list."),
        ("Phase 7: POSIX System Boundaries", "Days 31 - 35", "Configure zero-copy shared database mapper file index utilizing POSIX mmap."),
        ("Phase 8: Preprocessor Metaprogramming", "Days 36 - 40", "Generate dynamic Enum lists and translate functions using X-Macros tables."),
        ("Phase 9: Concurrency & MISRA C", "Days 41 - 45", "Create thread-safe Producer-Consumer ring buffer using POSIX mutexes and cond variables.")
    ]

    for idx, row_data in enumerate(gate_checklists, 24):
        ws_intro.row_dimensions[idx].height = 22
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_intro.cell(row=idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = border_thin
            if col_idx in [1, 3]:
                cell.alignment = align_left
            else:
                cell.alignment = align_center
        if idx % 2 == 1:
            for c in range(1, 4):
                ws_intro.cell(row=idx, column=c).fill = fill_zebra

    # Adjust Intro Column Widths
    ws_intro.column_dimensions["A"].width = 32
    ws_intro.column_dimensions["B"].width = 16
    ws_intro.column_dimensions["C"].width = 95

    # -------------------------------------------------------------
    # Populate Sheet 2: Daily Progress Tracker
    # -------------------------------------------------------------
    # Title Block
    ws_tasks.merge_cells("A1:I2")
    title_cell = ws_tasks.cell(row=1, column=1, value="C SYSTEMS PROGRAMMING ELITE 45-DAY FAST-TRACK")
    title_cell.font = font_title
    title_cell.alignment = align_center

    headers = [
        "Day", "Phase", "Core Focus Topic", "Vetting Challenge Task", 
        "Target Date", "Status", "Weight", "Grade / Audit", "Notes / Blockages"
    ]
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws_tasks.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = align_center
        cell.border = border_thin
    
    ws_tasks.row_dimensions[4].height = 28

    # Populate Data
    row_idx = 5
    day_counter = 0
    for phase_name, days_list in phases:
        for topic, challenge in days_list:
            ws_tasks.row_dimensions[row_idx].height = 20
            
            # Write columns
            ws_tasks.cell(row=row_idx, column=1, value=f"Day {day_counter+1:02d}").font = font_body_bold
            ws_tasks.cell(row=row_idx, column=1).alignment = align_center
            
            ws_tasks.cell(row=row_idx, column=2, value=phase_name).font = font_body
            ws_tasks.cell(row=row_idx, column=2).alignment = align_left
            
            ws_tasks.cell(row=row_idx, column=3, value=topic).font = font_body
            ws_tasks.cell(row=row_idx, column=3).alignment = align_left
            
            ws_tasks.cell(row=row_idx, column=4, value=challenge).font = font_body
            ws_tasks.cell(row=row_idx, column=4).alignment = align_left
            
            ws_tasks.cell(row=row_idx, column=5, value=dates[day_counter]).font = font_body
            ws_tasks.cell(row=row_idx, column=5).alignment = align_center
            
            # Status defaults to "Not Started"
            status_cell = ws_tasks.cell(row=row_idx, column=6, value="Not Started")
            status_cell.font = font_body_bold
            status_cell.alignment = align_center
            
            # Weight is exactly 1/45
            weight_cell = ws_tasks.cell(row=row_idx, column=7, value=1/45)
            weight_cell.font = font_body
            weight_cell.alignment = align_right
            weight_cell.number_format = "0.0%"
            
            # Grade defaults to "Pending"
            grade_cell = ws_tasks.cell(row=row_idx, column=8, value="Pending")
            grade_cell.font = font_body
            grade_cell.alignment = align_center
            
            # Notes / Blockages
            ws_tasks.cell(row=row_idx, column=9, value="").alignment = align_left
            
            # Apply border to all cells in the row
            for c in range(1, 10):
                ws_tasks.cell(row=row_idx, column=c).border = border_thin
                
            # Zebra striping
            if row_idx % 2 == 0:
                for c in range(1, 10):
                    ws_tasks.cell(row=row_idx, column=c).fill = fill_zebra
            
            row_idx += 1
            day_counter += 1

    # Add data validation to Status and Grade columns
    from openpyxl.worksheet.datavalidation import DataValidation
    dv_status = DataValidation(type="list", formula1='"Not Started,In Progress,Completed"', allow_blank=True)
    dv_grade = DataValidation(type="list", formula1='"Pending,A+ [100%],A [90%],B [80%],C [70%],F [Fail]"', allow_blank=True)
    
    ws_tasks.add_data_validation(dv_status)
    ws_tasks.add_data_validation(dv_grade)
    
    dv_status.add(f"F5:F49")
    dv_grade.add(f"H5:H49")

    # Auto-adjust column widths
    for col in ws_tasks.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2]: continue # Skip merged header
            if cell.value:
                # Check formatted values or formulas length
                max_len = max(max_len, len(str(cell.value)))
        ws_tasks.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Adjust special columns widths
    ws_tasks.column_dimensions["A"].width = 10
    ws_tasks.column_dimensions["B"].width = 30
    ws_tasks.column_dimensions["C"].width = 45
    ws_tasks.column_dimensions["D"].width = 50
    ws_tasks.column_dimensions["E"].width = 14
    ws_tasks.column_dimensions["F"].width = 14
    ws_tasks.column_dimensions["G"].width = 10
    ws_tasks.column_dimensions["H"].width = 14
    ws_tasks.column_dimensions["I"].width = 30

    # -------------------------------------------------------------
    # Populate Sheet 1: Executive Dashboard
    # -------------------------------------------------------------
    # Title Block
    ws_dash.merge_cells("A1:G2")
    dash_title = ws_dash.cell(row=1, column=1, value="C SYSTEMS PROGRAMMING EXECUTIVE DASHBOARD")
    dash_title.font = font_title
    dash_title.alignment = align_center

    # 1. KPI Blocks Header
    def setup_kpi_card(ws, label, formula, col_start, col_end):
        ws.merge_cells(f"{col_start}4:{col_end}4")
        lbl_cell = ws[f"{col_start}4"]
        lbl_cell.value = label
        lbl_cell.font = font_kpi_lbl
        lbl_cell.alignment = align_center
        lbl_cell.fill = fill_kpi
        
        ws.merge_cells(f"{col_start}5:{col_end}5")
        val_cell = ws[f"{col_start}5"]
        val_cell.value = formula
        val_cell.font = font_kpi_num
        val_cell.alignment = align_center
        val_cell.fill = fill_kpi
        
        # Apply borders to card area
        for r in [4, 5]:
            for c in range(openpyxl.utils.column_index_from_string(col_start), openpyxl.utils.column_index_from_string(col_end) + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border_kpi

    setup_kpi_card(ws_dash, "TOTAL TARGET DAYS", 45, "A", "B")
    setup_kpi_card(ws_dash, "COMPLETED", "=COUNTIF('Daily Progress Tracker'!F5:F49, \"Completed\")", "C", "C")
    setup_kpi_card(ws_dash, "IN PROGRESS", "=COUNTIF('Daily Progress Tracker'!F5:F49, \"In Progress\")", "D", "D")
    setup_kpi_card(ws_dash, "NOT STARTED", "=COUNTIF('Daily Progress Tracker'!F5:F49, \"Not Started\")", "E", "E")
    
    # Overall Completion Percentage Card
    ws_dash.merge_cells("F4:G4")
    lbl_comp = ws_dash.cell(row=4, column=6, value="OVERALL COMPLETION %")
    lbl_comp.font = font_kpi_lbl
    lbl_comp.alignment = align_center
    lbl_comp.fill = fill_kpi
    lbl_comp.border = border_kpi
    
    ws_dash.merge_cells("F5:G5")
    val_comp = ws_dash.cell(row=5, column=6, value="=C5/A5")
    val_comp.font = font_kpi_num
    val_comp.alignment = align_center
    val_comp.fill = fill_kpi
    val_comp.number_format = "0.0%"
    val_comp.border = border_kpi

    ws_dash.row_dimensions[4].height = 18
    ws_dash.row_dimensions[5].height = 32

    # 2. Phase-wise Summary Table
    ws_dash.cell(row=8, column=1, value="TRAINING PHASE PERFORMANCE HIGHLIGHTS").font = font_section
    
    summary_headers = ["Phase / Module Name", "Days Included", "Status", "Completion %", "Milestone Target Gate"]
    for col_idx, sh in enumerate(summary_headers, 1):
        cell = ws_dash.cell(row=9, column=col_idx, value=sh)
        cell.font = font_header
        cell.fill = fill_accent
        cell.alignment = align_center
        cell.border = border_thin
    
    ws_dash.row_dimensions[9].height = 24

    # Formulas map directly to the 5-day sections on Sheet 2
    phases_summary = [
        ("Phase 1: Compiler & Toolchains", "Day 01 - 05", 
         "=IF(COUNTIF('Daily Progress Tracker'!F5:F9, \"Completed\")=5, \"Completed\", IF(COUNTIF('Daily Progress Tracker'!F5:F9, \"Not Started\")=5, \"Not Started\", \"In Progress\"))", 
         "=COUNTIF('Daily Progress Tracker'!F5:F9, \"Completed\")/5", "Makefile Dependency Hook"),
        
        ("Phase 2: Data & Radix Hazards", "Day 06 - 10", 
         "=IF(COUNTIF('Daily Progress Tracker'!F10:F14, \"Completed\")=5, \"Completed\", IF(COUNTIF('Daily Progress Tracker'!F10:F14, \"Not Started\")=5, \"Not Started\", \"In Progress\"))", 
         "=COUNTIF('Daily Progress Tracker'!F10:F14, \"Completed\")/5", "Safe Float-Compare Parser"),
        
        ("Phase 3: Stack & Execution Frame", "Day 11 - 15", 
         "=IF(COUNTIF('Daily Progress Tracker'!F15:F19, \"Completed\")=5, \"Completed\", IF(COUNTIF('Daily Progress Tracker'!F15:F19, \"Not Started\")=5, \"Not Started\", \"In Progress\"))", 
         "=COUNTIF('Daily Progress Tracker'!F15:F19, \"Completed\")/5", "Stack Canary Overflow Audit"),
        
        ("Phase 4: Pointers & Arithmetic", "Day 16 - 20", 
         "=IF(COUNTIF('Daily Progress Tracker'!F20:F24, \"Completed\")=5, \"Completed\", IF(COUNTIF('Daily Progress Tracker'!F20:F24, \"Not Started\")=5, \"Not Started\", \"In Progress\"))", 
         "=COUNTIF('Daily Progress Tracker'!F20:F24, \"Completed\")/5", "Jump-Table Callbacks Map"),
        
        ("Phase 5: Structures & Alignments", "Day 21 - 25", 
         "=IF(COUNTIF('Daily Progress Tracker'!F25:F29, \"Completed\")=5, \"Completed\", IF(COUNTIF('Daily Progress Tracker'!F25:F29, \"Not Started\")=5, \"Not Started\", \"In Progress\"))", 
         "=COUNTIF('Daily Progress Tracker'!F25:F29, \"Completed\")/5", "Serialized Pack/Unpack"),
        
        ("Phase 6: Custom Allocators", "Day 26 - 30", 
         "=IF(COUNTIF('Daily Progress Tracker'!F30:F34, \"Completed\")=5, \"Completed\", IF(COUNTIF('Daily Progress Tracker'!F30:F34, \"Not Started\")=5, \"Not Started\", \"In Progress\"))", 
         "=COUNTIF('Daily Progress Tracker'!F30:F34, \"Completed\")/5", "O(1) Slab Allocator"),
        
        ("Phase 7: POSIX System Boundaries", "Day 31 - 35", 
         "=IF(COUNTIF('Daily Progress Tracker'!F35:F39, \"Completed\")=5, \"Completed\", IF(COUNTIF('Daily Progress Tracker'!F35:F39, \"Not Started\")=5, \"Not Started\", \"In Progress\"))", 
         "=COUNTIF('Daily Progress Tracker'!F35:F39, \"Completed\")/5", "Memory-Mapped Database Engine"),
        
        ("Phase 8: Preprocessor Metaprogramming", "Day 36 - 40", 
         "=IF(COUNTIF('Daily Progress Tracker'!F40:F44, \"Completed\")=5, \"Completed\", IF(COUNTIF('Daily Progress Tracker'!F40:F44, \"Not Started\")=5, \"Not Started\", \"In Progress\"))", 
         "=COUNTIF('Daily Progress Tracker'!F40:F44, \"Completed\")/5", "X-Macros Dynamic Lookup Table"),
         
        ("Phase 9: Concurrency & MISRA C", "Day 41 - 45", 
         "=IF(COUNTIF('Daily Progress Tracker'!F45:F49, \"Completed\")=5, \"Completed\", IF(COUNTIF('Daily Progress Tracker'!F45:F49, \"Not Started\")=5, \"Not Started\", \"In Progress\"))", 
         "=COUNTIF('Daily Progress Tracker'!F45:F49, \"Completed\")/5", "MISRA pthreads Scheduler")
    ]

    for idx, row_data in enumerate(phases_summary, 10):
        ws_dash.row_dimensions[idx].height = 22
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_dash.cell(row=idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = border_thin
            
            if col_idx in [1, 5]:
                cell.alignment = align_left
            elif col_idx in [2, 3]:
                cell.alignment = align_center
            elif col_idx == 4:
                cell.alignment = align_right
                cell.number_format = "0%"

            if col_idx == 3:
                cell.font = font_body_bold

        if idx % 2 == 1:
            for c in range(1, 6):
                ws_dash.cell(row=idx, column=c).fill = fill_zebra

    # Add overall status totals row
    tot_row = 19
    ws_dash.cell(row=tot_row, column=1, value="Overall Program Status").font = font_body_bold
    ws_dash.cell(row=tot_row, column=1).alignment = align_left
    ws_dash.cell(row=tot_row, column=1).border = double_bottom
    
    ws_dash.cell(row=tot_row, column=2, value="45 Days").font = font_body_bold
    ws_dash.cell(row=tot_row, column=2).alignment = align_center
    ws_dash.cell(row=tot_row, column=2).border = double_bottom

    ws_dash.cell(row=tot_row, column=3, value="=IF(F5=1, \"Completed\", \"Active\")").font = font_body_bold
    ws_dash.cell(row=tot_row, column=3).alignment = align_center
    ws_dash.cell(row=tot_row, column=3).border = double_bottom
    
    ws_dash.cell(row=tot_row, column=4, value="=AVERAGE(D10:D18)").font = font_body_bold
    ws_dash.cell(row=tot_row, column=4).alignment = align_right
    ws_dash.cell(row=tot_row, column=4).number_format = "0%"
    ws_dash.cell(row=tot_row, column=4).border = double_bottom
    
    ws_dash.cell(row=tot_row, column=5, value="Final Graduation").font = font_body_bold
    ws_dash.cell(row=tot_row, column=5).alignment = align_left
    ws_dash.cell(row=tot_row, column=5).border = double_bottom

    # Adjust Dashboard Widths
    ws_dash.column_dimensions["A"].width = 38
    ws_dash.column_dimensions["B"].width = 18
    ws_dash.column_dimensions["C"].width = 18
    ws_dash.column_dimensions["D"].width = 16
    ws_dash.column_dimensions["E"].width = 38
    ws_dash.column_dimensions["F"].width = 14
    ws_dash.column_dimensions["G"].width = 14

    # -------------------------------------------------------------
    # Dynamic Openpyxl Column Progress Chart
    # -------------------------------------------------------------
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Milestone Phase Completion Progress"
    chart.y_axis.title = "Completion %"
    chart.x_axis.title = "Phases"
    chart.y_axis.scaling.min = 0.0
    chart.y_axis.scaling.max = 1.0
    chart.height = 14
    chart.width = 22

    # Columns: Completion % (Col D, row 9 to 18)
    data = Reference(ws_dash, min_col=4, min_row=9, max_row=18)
    # Columns: Phase / Module Name (Col A, row 10 to 18)
    cats = Reference(ws_dash, min_col=1, min_row=10, max_row=18)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None # Single series, no legend needed

    # Add chart below the summary table at A22
    ws_dash.add_chart(chart, "A22")

    # -------------------------------------------------------------
    # Conditional Formatting Rules
    # -------------------------------------------------------------
    # Colors
    fill_green = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Completed
    font_green = Font(name=FONT_FAMILY, size=10, bold=True, color="166534")
    
    fill_yellow = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid") # In Progress
    font_yellow = Font(name=FONT_FAMILY, size=10, bold=True, color="854D0E")

    fill_gray = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid") # Not Started
    font_gray = Font(name=FONT_FAMILY, size=10, bold=True, color="475569")

    # Apply to Tasks Tracker Status Column (F5:F49)
    ws_tasks.conditional_formatting.add("F5:F49", CellIsRule(operator="equal", formula=['"Completed"'], fill=fill_green, font=font_green))
    ws_tasks.conditional_formatting.add("F5:F49", CellIsRule(operator="equal", formula=['"In Progress"'], fill=fill_yellow, font=font_yellow))
    ws_tasks.conditional_formatting.add("F5:F49", CellIsRule(operator="equal", formula=['"Not Started"'], fill=fill_gray, font=font_gray))

    # Apply to Dashboard Status Column (C10:C18)
    ws_dash.conditional_formatting.add("C10:C18", CellIsRule(operator="equal", formula=['"Completed"'], fill=fill_green, font=font_green))
    ws_dash.conditional_formatting.add("C10:C18", CellIsRule(operator="equal", formula=['"In Progress"'], fill=fill_yellow, font=font_yellow))
    ws_dash.conditional_formatting.add("C10:C18", CellIsRule(operator="equal", formula=['"Not Started"'], fill=fill_gray, font=font_gray))

    # Save Workbook
    out_path = os.path.join(C_FAST_TRACK_DIR, "c_fast_track_tracker.xlsx")
    wb.save(out_path)
    print(f"SUCCESS: Elite C Fast-Track Excel progress tracker compiled at: {out_path}")

if __name__ == "__main__":
    main()
