#!/usr/bin/env python3
import os
import sys
import argparse
import subprocess
import shutil
import datetime

# Color Codes for Premium Console Outputs
GREEN = "\033[92m"
YELLOW = "\033[93m"
RED = "\033[91m"
BLUE = "\033[94m"
CYAN = "\033[96m"
BOLD = "\033[1m"
RESET = "\033[0m"

# Absolute Workspace Configuration
WORKSPACE = "/home/siva/sivaramireddy/Project1"
C_FAST_TRACK_DIR = os.path.join(WORKSPACE, "c-fast-track")
TRACKER_PATH = os.path.join(C_FAST_TRACK_DIR, "c_fast_track_tracker.xlsx")

def log_info(msg):
    print(f"{BLUE}[INFO]{RESET} {msg}")

def log_success(msg):
    print(f"{GREEN}[PASS]{RESET} {BOLD}{msg}{RESET}")

def log_warn(msg):
    print(f"{YELLOW}[WARN]{RESET} {msg}")

def log_error(msg):
    print(f"{RED}[FAIL]{RESET} {BOLD}{msg}{RESET}")

def run_cmd(cmd, description):
    log_info(f"Running: {description}...")
    try:
        result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, check=True)
        return True, result.stdout, result.stderr
    except subprocess.CalledProcessError as e:
        return False, e.stdout, e.stderr

def check_dependencies():
    deps = ["gcc", "valgrind", "cppcheck"]
    log_info("Auditing toolchain dependencies in system path...")
    for dep in deps:
        path = shutil.which(dep)
        if path:
            print(f"  - {dep:<10}: {GREEN}Available{RESET} ({path})")
        else:
            print(f"  - {dep:<10}: {YELLOW}Missing{RESET}")

def compile_c_code(file_path):
    output_bin = "./temp_autovet_bin"
    # Strict industrial system flags
    gcc_flags = [
        "gcc", "-std=c11", "-Wall", "-Wextra", "-Werror", "-pedantic", 
        "-g3", "-O0", file_path, "-o", output_bin, "-pthread", "-lm"
    ]
    
    success, stdout, stderr = run_cmd(gcc_flags, "GCC compilation with strict diagnostic flags")
    if not success:
        log_error("Compilation aborted due to compiler errors / warnings.")
        print(f"{RED}{stderr}{RESET}")
        return False, None
    
    log_success("Code compiled cleanly with ZERO compiler warnings/errors under pedantic C11!")
    return True, output_bin

def run_static_analysis(file_path):
    if not shutil.which("cppcheck"):
        log_warn("cppcheck not available in path. Skipping static analysis audit.")
        return True
    
    cppcheck_cmd = ["cppcheck", "--enable=all", "--error-exitcode=1", "--std=c11", file_path]
    success, stdout, stderr = run_cmd(cppcheck_cmd, "cppcheck static safety checks")
    if not success:
        log_error("Static analysis failed. MISRA or memory violation risks detected!")
        print(f"{RED}{stderr}{RESET}")
        return False
    
    log_success("Static analysis audit passed! Zero safety violations detected.")
    return True

def run_memory_analysis(binary_path):
    if not shutil.which("valgrind"):
        log_warn("Valgrind not available. Skipping dynamic memory leak checks.")
        # Execute binary directly
        success, stdout, stderr = run_cmd([binary_path], "raw binary execution")
        if not success:
            log_error("Program execution crashed with runtime error.")
            print(f"{RED}{stderr}{RESET}")
            return False
        log_success("Binary executed successfully without crashes.")
        return True

    valgrind_cmd = [
        "valgrind", "--leak-check=full", "--error-exitcode=1", 
        "--show-leak-kinds=all", binary_path
    ]
    success, stdout, stderr = run_cmd(valgrind_cmd, "Valgrind dynamic memory tracking")
    if not success:
        log_error("Dynamic memory check failed! Leaks or uninitialized read detected.")
        print(f"{RED}{stderr}{RESET}")
        return False
    
    log_success("Dynamic memory checks passed! Zero memory leaks or page faults detected.")
    return True

def update_tracker(day):
    if not os.path.exists(TRACKER_PATH):
        log_warn(f"Progress tracker workbook not found at: {TRACKER_PATH}. Skipping status updates.")
        return
    
    try:
        import openpyxl
    except ImportError:
        log_warn("openpyxl library not installed. Cannot update spreadsheet progress.")
        return

    try:
        wb = openpyxl.load_workbook(TRACKER_PATH)
        ws_tasks = wb["Daily Progress Tracker"]
        
        # Day row index calculation: Day 1 is row 5, Day 2 is row 6... Day 45 is row 49
        row_num = day + 4
        
        # Audit sheet boundary check
        day_cell = ws_tasks.cell(row=row_num, column=1)
        if day_cell.value != f"Day {day:02d}":
            log_warn(f"Mismatch in sheet row mapping. Expected Row {row_num} to be 'Day {day:02d}' but got '{day_cell.value}'. Search triggered...")
            # Fallback scan
            matched = False
            for r in range(5, 55):
                if ws_tasks.cell(row=r, column=1).value == f"Day {day:02d}":
                    row_num = r
                    matched = True
                    break
            if not matched:
                log_error(f"Could not locate 'Day {day:02d}' inside Daily Progress Tracker sheet.")
                return

        # Update columns F (Status), H (Grade), I (Notes)
        ws_tasks.cell(row=row_num, column=6, value="Completed") # Status
        ws_tasks.cell(row=row_num, column=8, value="A+ [100%]")   # Grade
        
        current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
        ws_tasks.cell(row=row_num, column=9, value=f"Passed Autovet Harness on {current_time}") # Notes
        
        wb.save(TRACKER_PATH)
        log_success(f"Spreadsheet Tracker updated! Row {row_num} [Day {day:02d}] marked as COMPLETED with A+ grade.")
        
    except Exception as e:
        log_error(f"Error occurred while updating workbook: {e}")

def main():
    parser = argparse.ArgumentParser(description="Minux Academy Elite 45-Day C Systems Autovet Harness")
    parser.add_argument("--day", type=int, required=True, help="The target Day number to vet (1-45)")
    parser.add_argument("--file", type=str, required=True, help="Path to your C source code solution file")
    args = parser.parse_args()

    print(f"\n{BOLD}{CYAN}===================================================================={RESET}")
    print(f"  {BOLD}MINUX ACADEMY: C SYSTEMS PROGRAMMING AUTOVET HARNESS (DAY {args.day:02d}){RESET}")
    print(f"{BOLD}{CYAN}===================================================================={RESET}\n")

    if args.day < 1 or args.day > 45:
        log_error("Invalid day target. Please select a day between 1 and 45.")
        sys.exit(1)

    if not os.path.exists(args.file):
        log_error(f"Source file not found: {args.file}")
        sys.exit(1)

    check_dependencies()
    print("-" * 68)

    # Stage 1: Static Analysis Audit
    static_pass = run_static_analysis(args.file)
    if not static_pass:
        log_error(f"Day {args.day:02d} Vetting aborted. Fix static analysis violations.")
        sys.exit(1)

    # Stage 2: Compilation Audit
    compile_pass, binary_path = compile_c_code(args.file)
    if not compile_pass:
        log_error(f"Day {args.day:02d} Vetting aborted. Fix compiler warnings/errors.")
        sys.exit(1)

    # Stage 3: Dynamic Memory Audit
    mem_pass = run_memory_analysis(binary_path)
    
    # Cleanup temp binary
    if binary_path and os.path.exists(binary_path):
        os.remove(binary_path)

    if not mem_pass:
        log_error(f"Day {args.day:02d} Vetting aborted. Fix memory leaks or runtime bugs.")
        sys.exit(1)

    # Stage 4: Record Success in Progress Tracker
    print("-" * 68)
    log_success(f"CONGRATULATIONS! Day {args.day:02d} Vetting completed successfully with elite marks!")
    update_tracker(args.day)
    print(f"\n{BOLD}{CYAN}===================================================================={RESET}\n")

if __name__ == "__main__":
    main()
